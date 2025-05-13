import asyncio
import json
import logging
import os
import random
import re
import regex
import shutil
import sys
import time
import traceback
import zipfile
from datetime import datetime
from io import BytesIO
from typing import List
from urllib.parse import urlparse, unquote

import pandas as pd
from bs4 import BeautifulSoup, Tag, NavigableString
from copy import copy
# from fuzzywuzzy import fuzz
from rapidfuzz import fuzz
from fastapi import FastAPI, Request, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.responses import JSONResponse
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from googletrans import Translator
from html import unescape
from html2text import html2text
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from PIL import Image, UnidentifiedImageError
from playsound import playsound
from pydantic import BaseModel
from starlette.staticfiles import StaticFiles

from data_processing import convert_data_to_files
from demo import period_check
from parser import extract_content, Https_Parser
from utils import load_config, blacklist, create_zip_archive, add_unreachable_site, get_google_search_results_alt, download_images_v2, gen_rand_text, claimed_names # get_google_search_results
from html_to_docx import generate_document

from selenium_stealth import stealth
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager



from config_chrome_options import chrome_options

import duplicate_clear

# Налаштування статичних файлів
static_dir = 'templates'
os.makedirs(static_dir, exist_ok=True)
os.makedirs('static', exist_ok=True)
os.makedirs('templates/static', exist_ok=True)

app = FastAPI()
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="templates/static"), name="static")
# Налаштування логування
logging.basicConfig(level=logging.INFO, filename='parser.log', filemode='a',
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Завантаження конфігурації
config = load_config('config.yaml')

# Глобальна змінна для зберігання результатів парсингу
parsed_data = pd.DataFrame()
translator = Translator()


@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    """
    Головна сторінка сайту.
    period_check перевірка на демо період
    """
    return period_check(templates.TemplateResponse("index.html", {"request": request}))


@app.post("/parse")
async def parse_url(request: Request, proxy_chrome=None):
    """
    Парсинг URL або списку URL.

    Args:
        request (Request): HTTP запит з URL або списком URL.

    Returns:
        HTMLResponse: HTML-сторінка з результатами парсингу або повідомленням про відсутність даних.
    """
    global parsed_data
    parsed_data = pd.DataFrame()  # Збір результатів гугл пошуку
    # ЗВУК ПРИ ПОЧАТКУ ПАРСИНГУ
    playsound(r'./audio/rozpochala_parsing.mp3')

    form = await request.form()
    print(form)
    from_save = form.get('fromSave') == 'true'
    url = form.get('url')
    urls = [form.get(f'urls_{i}').splitlines() for i in range(len(form)) if form.get(f'urls_{i}')]
    code_v = form.get('code_v', '0')  # Значення за замовчуванням
    parser_type = form.get('parser_type', 'https')  # Значення за замовчуванням
    if parser_type == "Selenium":
        parser_type = "Selenium-old"
    min_chars = int(form.get('min_chars', 0))  # Значення за замовчуванням
    max_chars = int(form.get('max_chars', -1))  # Значення за замовчуванням
    break_list = config.get('break_words', [])
    ignore_words = config.get('ignore_words', [])
    ignore_sentence = config.get('ignore_sentence', [])
    toggleButtonFilterTop = form.get('toggleButtonFilterTop') == 'true' # Виключення топ фільтра включити true, виключити false
    toggleButtonFilterImage = form.get('toggleButtonFilterImage') == 'true' # Виключення фільтра картинки, аналогічно.

    def yield_ID(start_num=2):
        """
        Генератор для створення унікальних ID.

        Args:
            start_num (int): Початковий номер для генерації ID.

        Yields:
            str: Унікальний ID.
        """
        num = start_num
        while True:
            ID = f'1.{num}'
            yield ID
            num += 1

    def block(url):
        black_list = blacklist('blacklist.txt')
        for url_stop in black_list:
            if url_stop == url:
                logging.info(f'Сайт занесений в blacklist\t\t {url}')
                print(f'INFO: Сайт занесений в blacklist\t\t {url}')
                return False
        return True

    def limit_text(text: BeautifulSoup, min_chars: int, max_chars: int, url: str):
        text_all = ' '.join(text.stripped_strings)
        if len(text_all) < min_chars:
            logging.info(f'Кількість символів занадто мала\t\t{url}')
            print(f'INFO: Кількість символів занадто мала\t\t{url}')
            return False
        if max_chars != -1 and len(text_all) > max_chars:
            logging.info(f'Кількість символів занадто велика\t\t{url}')
            print(f'INFO: Кількість символів занадто велика\t\t{url}')
            return False
        return True

    id_generator = yield_ID()
    all_data = []

    def log_unreachable_sites(data):
        if data['Status Parsing'] == 'НІ':
            try:
                domain = data["URL"].split("/")[2]
            except:
                domain = data["URL"]
            add_unreachable_site('Blacklist_Domen.txt', domain)
            print(f'INFO: Домен доданий до Blacklist_Domen.txt\t\t {domain}')

    def log_ok_parser(data):
        if data['Status Parsing'] == 'ТАК':
            domain = data["URL"]
            add_unreachable_site('Blacklist_Page.txt', domain)
            print(f'INFO: URL доданий до Blacklist_Page.txt\t\t {domain}')

    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options())
    if url:
        if block(url):
            data = await extract_content(url,
                                         break_list,
                                         code_v=code_v,
                                         parser_type=parser_type,
                                         ignore_words=ignore_words,
                                         ignore_sentence=ignore_sentence)
            if data and limit_text(data['Content'], min_chars, max_chars, url):
                data['ID'] = next(id_generator)
                log_ok_parser(data)
                all_data.append(data)
            else:
                logging.error('Не вдалося отримати дані з URL або зміст сайту недійсний')
                print('ERROR: Не вдалося отримати дані з URL або зміст сайту недійсний')
                log_unreachable_sites(data)
                return period_check(
                    HTMLResponse(content="<h1>Не вдалося отримати дані з URL або зміст сайту недійсний</h1>",
                                 status_code=404))
        else:
            return HTMLResponse(content="<h1>Сайт занесений в blacklist</h1>", status_code=404)
    elif urls or from_save:
        if from_save:
            try:
                with open("static/last_parse_save.json", "r", encoding="utf-8") as f:
                    loaded_save = json.load(f)
                print("Loaded data from json save file")
            except:
                print("Can not load json save file")
                return HTMLResponse(content="Error: Can not load json save file", status_code=400)

            last_index = loaded_save.get("last_index", [])
            last_generator_id = loaded_save.get("last_generator_id", [])
            urls = loaded_save.get("urls", [])
            code_v = loaded_save.get("code_v", [])
            parser_type = loaded_save.get("parser_type", [])
            min_chars = loaded_save.get("min_chars", [])
            max_chars = loaded_save.get("max_chars", [])
            break_list = loaded_save.get("break_list", [])
            ignore_words = loaded_save.get("ignore_words", [])
            ignore_sentence = loaded_save.get("ignore_sentence", [])
            toggleButtonFilterTop = loaded_save.get("toggleButtonFilterTop", [])
            toggleButtonFilterImage = loaded_save.get("toggleButtonFilterImage", [])
            all_data = loaded_save.get("all_data", [])

            id_generator = yield_ID(int(last_generator_id.split(".")[1]) + 1)
        else:
            last_index = -1

        for ix, url_block in enumerate(urls):
            if ix <= last_index:
                continue

            tasks = []
            url_block = [url for url in url_block if block(url)]
            for url in url_block:
                url = url.strip()
                tasks.append(extract_content(url,
                                             break_list,
                                             code_v=code_v,
                                             parser_type=parser_type,
                                             ignore_words=ignore_words,
                                             ignore_sentence=ignore_sentence))
            results = await asyncio.gather(*tasks)
            results_new = []

            # Множина для збереження унікального контенту
            unique_content = set()
            unique_results = []
            for item in results:
                content_text = item['Content'].text.strip()
                if content_text not in unique_content:
                    unique_content.add(content_text)
                    unique_results.append(item)
            results = unique_results

            if toggleButtonFilterImage:
                for data in results:
                    """Видалити колонки, в яких немає зображень (колонка 'Image now Url' пуста)"""
                    if data['Image Url_original'] == '':
                        continue  # Пропустити цей запис
                    results_new.append(data)
                results = results_new

            if toggleButtonFilterTop:
                filterInput = int(form.get('filterInput', 5))
                """Зробити сортування по кількості символів в Content"""
                # Сортуємо список, очищуючи текст від зайвих пробілів
                sorted_data = sorted(results, key=lambda x: len(x['Content'].text.strip()), reverse=True)

                # Перевірка результату
                print([len(data['Content'].text.strip()) for data in sorted_data])
                results = sorted_data[:filterInput]

            last_generator_id = 0
            for data in results:
                if data and limit_text(data['Content'], min_chars, max_chars, url):
                    log_ok_parser(data)
                    last_generator_id = next(id_generator)
                    data['ID'] = last_generator_id
                    all_data.append(data)
                else:
                    log_unreachable_sites(data)
                    logging.error('Не вдалося отримати дані з URL або зміст сайту недійсний')
                    print('ERROR: Не вдалося отримати дані з URL або зміст сайту недійсний')

            if ix % 1 == 0:
                print(f"Парсинг збереженно в тимчасовий файл.. (index: {ix})")

                dumped_save = dict()
                dumped_save["last_index"] = ix
                dumped_save["last_generator_id"] = last_generator_id
                dumped_save["urls"] = urls
                dumped_save["code_v"] = code_v
                dumped_save["parser_type"] = parser_type
                dumped_save["min_chars"] = min_chars
                dumped_save["max_chars"] = max_chars
                dumped_save["break_list"] = break_list
                dumped_save["ignore_words"] = ignore_words
                dumped_save["ignore_sentence"] = ignore_sentence
                dumped_save["toggleButtonFilterTop"] = toggleButtonFilterTop
                dumped_save["toggleButtonFilterImage"] = toggleButtonFilterImage
                dumped_save["all_data"] = all_data

                def default(obj):
                    if isinstance(obj, BeautifulSoup):
                        return unescape(str(obj))

                    return obj

                with open("static/last_parse_save.json", "w", encoding="utf-8") as f:
                    json.dump(dumped_save, f, ensure_ascii=False, indent=4, default=default)


    # Удаляє дубльований Url в Exel. В екселі залищаються тільки унікальні Url
    if all_data:
        parsed_data = pd.DataFrame(all_data)
        parsed_data = parsed_data.drop_duplicates(subset=['URL'])
        id_generator_new = yield_ID()
        parsed_data['ID'] = [next(id_generator_new) for _ in range(len(parsed_data))]

    if parsed_data is not None and not parsed_data.empty:
        # ЗВУК ПРИ ЗАВЕРШЕННІ ПАРСИНГУ
        try:
            print('audio')
            playsound(r'./audio/zaversheno_parsing.mp3')
        except Exception as e:
            print(e)
        return period_check(templates.TemplateResponse("parsed_result.html", {"request": request}))
    else:
        return HTMLResponse(content="<h1>No data available</h1>", status_code=404)


@app.get("/table", response_class=HTMLResponse)
async def display_table(request: Request):
    """
    Відображення таблиці з парсингованими даними.
    """
    if parsed_data is not None:
        parsed_data_table_view = parsed_data
        if config.get('cleaned_data_table_view', False):
            parsed_data_table_view = (
                parsed_data.apply(lambda x: x.map(lambda y: str(y).replace('\n', '').replace('\r', ' '))))
        html_table = parsed_data_table_view.to_html(index=False, border=1, classes='data-table')
        return period_check(
            templates.TemplateResponse("table_view.html", {"request": request, "html_table": html_table}))
    else:
        return HTMLResponse(content="<h1>No data available</h1>", status_code=404)


@app.get("/download")
async def download_file(filetype: str = "xlsx"):
    """
    Завантаження файлів у різних форматах (xlsx, csv, xml).
    """
    if parsed_data is None:
        return HTMLResponse(content="<h1>No data available</h1>", status_code=404)
    parsed_data_save = parsed_data
    if config.get('cleaned_data_save', False):
        parsed_data_save = (
            parsed_data.apply(lambda x: x.map(lambda y: str(y).replace('\n', '').replace('\r', ' '))))
    files_response = await convert_data_to_files(parsed_data_save, filetype)

    if isinstance(files_response, str):  # файл архіву
        return FileResponse(
            path=files_response,
            media_type='application/zip',
            filename=os.path.basename(files_response)
        )
    elif isinstance(files_response, tuple) and len(files_response) == 2:
        file_path, media_type = files_response
        if filetype == "xlsx":
            wb: Workbook = load_workbook(file_path)

            ws_0: Worksheet = wb._sheets[0]
            ws_0.title = "Статі"

            ws_1: Worksheet = wb.create_sheet(index=1)
            ws_1.title = "Переобход заголовків"
            ws_1.cell(1, 1, "ЗАГОЛОВКИ СТАРІ = 0 (НЕ ЗАБУТИ ДОДАТИ ЇХ В BLAKLIST)")
            ws_1.cell(1, 2, "НОВІ ЗАГОЛОВКИ (ПЕРЕВІРИТИ ЧАСТОТНІСТЬ)")
            ws_1.cell(1, 3, "Щоб удалити статю з не правильним заголовком (Ставити слово delete)")
            ws_1.cell(1, 1).font = Font(bold=True)
            ws_1.cell(1, 2).font = Font(bold=True)
            ws_1.cell(1, 3).font = Font(bold=True)
            ws_1.column_dimensions["A"].width = 60
            ws_1.column_dimensions["B"].width = 60
            ws_1.column_dimensions["C"].width = 60

            ws_2: Worksheet = wb.create_sheet(index=2)
            ws_2.title = "Фільтр по словах"
            ws_2.cell(1, 1, "STOP")

            ws_3: Worksheet = wb.create_sheet(index=3)
            ws_3.title = "Blacklist перевірка в гугл"

            ws_4: Worksheet = wb.create_sheet(index=4)
            ws_4.title = "ЗРОБИТИ ВРУЧНУ"

            ws_5: Worksheet = wb.create_sheet(index=5)
            ws_5.title = "Blacklist для дублів заголовків"

            wb.save(file_path)
        return FileResponse(
            path=file_path,
            media_type=media_type,
            filename=os.path.basename(file_path)
        )
    else:
        return period_check(HTMLResponse(content="<h1>Unsupported file type</h1>", status_code=400))


@app.post("/search", response_class=JSONResponse)
async def search_google(request: Request):
    """
    Пошук в Google за запитом.

    Args:
        request (Request): HTTP запит з параметром query.

    Returns:
        JSONResponse: JSON відповідь з результатами пошуку.
    """
    # ЗВУК початок парсингу гугл пошуку

    playsound(r'./audio/rozpochala_zbir_rezyltativ_google.mp3')
    data = await request.form()
    print(data)
    raw_queries = data.getlist('query[]')
    from_save = data.get('fromSave') == 'true'
    if from_save:
        try:
            with open("static/last_search_save.json", "r", encoding="utf-8") as f:
                loaded_save = json.load(f)
            print("Loaded data from json save file")
        except:
            print("Can not load json save file")
            return HTMLResponse(content="Error: Can not load json save file", status_code=400)

        queries = loaded_save.get("queries", [])
        num_results = loaded_save.get("num_results", [])
        last_index = loaded_save.get("last_index", [])
        list_search_google = loaded_save.get("list_search_google", [])

    else:
        queries = []
        for raw_query in raw_queries:
            split_queries = raw_query.split('\n')  # Розбиваємо на окремі запити
            cleaned_queries = [q.strip().replace('\r', '') for q in split_queries if q.strip()]  # Очищаємо кожен запит
            queries.extend(cleaned_queries)  # Додаємо очищені запити до загального списку

        num_results = data.get('num_results', 10)  # Значення за замовчуванням
        list_search_google = []
        last_index = -1


    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options())
    for ix, query in enumerate(queries):
        if ix <= last_index:
            continue

        if not query:
            return JSONResponse(content={"error": "Query parameter 'query' is required"}, status_code=400)
        # results = get_google_search_results(query, num_results=int(num_results))
        results = get_google_search_results_alt(query, num_results=int(num_results), driver=driver)
        print(results)
        list_search_google.append(results)

        print(f"Пошук збереженно в тимчасовий файл.. (index: {ix})")

        dumped_save = dict()
        dumped_save["queries"] = queries
        dumped_save["num_results"] = num_results
        dumped_save["last_index"] = ix
        dumped_save["list_search_google"] = list_search_google

        with open("static/last_search_save.json", "w", encoding="utf-8") as f:
            json.dump(dumped_save, f, ensure_ascii=False, indent=4)

    print(list_search_google)

    try:
        driver.close()
        driver.quit()
    except:
        pass

    try:
        # ЗВУК ПРИ ЗАКІНЧЕННІ ПОШУКУ
        playsound(r'./audio/zakincheno_zbir_site_googl.mp3')
        return period_check(JSONResponse(content={"results": list_search_google}))
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/downloadImage", response_class=HTMLResponse)
async def download_file(blocks: List):
    if os.path.exists('static/v2'):
        shutil.rmtree('static/v2')

    for block in blocks:
        imageUrl = [url.strip() for url in re.split(r'\s+', block['imageUrl']) if url.strip()]

        if block['folderName'] and block['folderName'].strip():
            try:
                safe_title = re.sub(r'[\/\\:*?"<>|\t\n]', '_', block['folderName']).strip()
                download_folder = f'static/v2/{safe_title}'
            except Exception as e:
                logging.error(f'Назва має некоректний символ: {e}')
                return HTMLResponse(content="<h1>Error: Invalid folder name</h1>", status_code=400)
        else:
            download_folder = 'static/v2'
        await download_images_v2(imageUrl, download_folder)
    zip_file_path = f'static/archive_{random.randint(1000, 9999)}'
    filename = shutil.make_archive(zip_file_path, 'zip', 'static/v2')
    return filename, None


class Block(BaseModel):
    folderName: str
    imageUrl: str


class RequestData(BaseModel):
    blocks: List[Block]


@app.post("/downloadPicture")
async def handle_form(request: Request, request_data: RequestData):
    data = await request.json()
    blocks = data.get('blocks', [])
    zip_file_path, error = await download_file(blocks)

    if error:
        return HTMLResponse(content=f"<h1>{error}</h1>", status_code=400)
    return FileResponse(
        path=zip_file_path,
        media_type='application/zip',
        filename=os.path.basename(zip_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(zip_file_path)}"'
        }
    )


@app.get("/VisualizationTranslation", response_class=HTMLResponse)
async def VisualizationTranslation(request: Request):
    return period_check(templates.TemplateResponse("VisualizationTranslation.html", {"request": request}))


uploaded_workbook = None


@app.post("/process_excel")
async def process_excel(file: UploadFile = File(...)):
    global uploaded_workbook
    # Завантажуємо файл Excel
    contents = await file.read()

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=BytesIO(contents))

    # Отримання даних з першого листа
    sheet = uploaded_workbook.active
    data = []

    # Отримуємо заголовки стовпців
    headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки

    # Індекси потрібних стовпців
    content_index = headers.index("Content")+1

    i_row = 2
    count_duplicate = 0
    duplicate = duplicate_clear.DuplicateClear()
    while i_row <= sheet.max_row:
        content = sheet.cell(i_row, content_index)  # Беремо тільки "Content"
        # Перевіряємо, чи є контент та на дублікат
        remove = None
        if content.value:
            print("Перевірка на дублікати", i_row, '/', sheet.max_row )
            remove = duplicate.test_duplicate(content.value)
            if remove:
                count_duplicate += 1
        if remove or not content.value:
            sheet.delete_rows(idx=content.row) # Видаляємо пусті або дублікати
            continue
        duplicate.add_content(content.value)
        data.append(str(content.value))
        i_row += 1

    # Перевірка наявності даних і повернення
    if not data:
        return JSONResponse(status_code=400, content={"error": "Файл пустий або статті уже є в бд."})

    return JSONResponse(content={"contents": data, "duplicate": count_duplicate})


@app.get("/test_excel")
async def test_excel():
    print(123)
    uploaded_workbook = load_workbook(filename=r"C:/Users/Пользователь/Downloads/archive_6357/table_9658.xlsx")

    sheet = uploaded_workbook._sheets[0]

    headers = [cell.value for cell in sheet[1]]

    content_index = headers.index("Content") + 1

    i_row = 2
    cnt = 0

    total_text = ""

    while i_row <= sheet.max_row:
        content = sheet.cell(i_row, content_index).value

        if cnt >= 1000:
            break

        if not content:
            cnt += 1
            continue

        total_text += content

        i_row += 1

    cnt = 0
    for x in Path("C:/Users/Пользователь/Downloads/archive_6357").iterdir():
        if str(x.name) not in total_text:
            print(x)
            cnt += 1

    print(456, cnt)

# Скачування/завантаження картинок. Початок коду. Перемістив в інший .exe
# Скачування/завантаження картинок. Початок коду. Перемістив в інший .exe

# @app.post("/process_excel_2")
# async def process_excel_2(request: Request):
#     data = await request.form()
#
#     if os.path.exists('static/results'):
#         shutil.rmtree('static/results')
#
#     if not os.path.exists('static/results'):
#         os.mkdir('static/results')
#
#     folder_name = data.get('folderName')
#     file = data.get('file')
#
#     if folder_name and folder_name.strip():
#         try:
#             safe_title = re.sub(r'[\/\\:*?"<>|\t\n]', '_', folder_name).strip()
#             download_folder = f'static/results/{safe_title}'
#
#             if not os.path.exists(f'static/results/{safe_title}'):
#                 os.mkdir(f'static/results/{safe_title}')
#         except Exception as e:
#             logging.error(f'Назва має некоректний символ: {e}')
#             return HTMLResponse(content="<h1>Error: Invalid folder name</h1>", status_code=400)
#     else:
#         if not os.path.exists('static/results/image'):
#             os.mkdir('static/results/image')
#         download_folder = 'static/results/image'
#
#     # Завантажуємо файл Excel
#     contents = await file.read()
#
#     # Спробуємо відкрити файл
#     uploaded_workbook = load_workbook(filename=BytesIO(contents))
#
#     # Отримання даних з першого листа
#     sheet = uploaded_workbook.active
#
#     # Отримуємо заголовки стовпців
#     headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки
#
#     # Індекси потрібних стовпців
#     content_index = headers.index("Content")+1
#     orig_urls_index = headers.index("Image Url_original")+1
#     image_now_urls_index = headers.index("Image now Url")+1
#
#     # https://sykni.fff/wp-content/uploads/2024/09/                                              bearded_european_man_in_casual_peach_isolated_excited_cheerful_holding_gift_box.jpg
#     # https://ulanude.royalthai.ru/upload/medialibrary/ef3/qqtyncldrg4wsrdj3pesahoyyx4ezt98/     bearded_european_man_in_casual_peach_isolated_excited_cheerful_holding_gift_box.jpg
#
#     i_row = 2
#     cnt = 0
#
#     # ЗВУК ПРИ ПОЧАТКУ ПАРСИНГУ ЩОБ РОЗКОМЕНТУВАТИ, ПОТРІБНО УДАЛИТИ #
#     # playsound(r'./audio/rozpochala_parsing.mp3')
#     print("Downloading started...")
#
#     import requests
#     session = requests.session()
#
#     while i_row <= sheet.max_row:
#         content = sheet.cell(i_row, content_index).value  # Беремо тільки "Content"
#         orig_urls = sheet.cell(i_row, orig_urls_index).value  # Беремо тільки "Image Url_original"
#
#         sheet.cell(i_row, image_now_urls_index, "")
#
#         if cnt >= 1000:
#             break
#
#         if not orig_urls:
#             cnt += 1
#             continue
#
#         new_links = []
#
#         print(f"Processing line {i_row}..")
#
#         soup = BeautifulSoup(content, "html.parser")
#
#         for url in list(set(orig_urls.split(" "))):
#             # print(url)
#             url = url.split("?")[0]
#             need_clear = False
#             pic_paths = []
#             if not "." in url[-6:]: # шука точку в останніх 6 символах ссилки
#                 need_clear = True
#             else:
#                 pic_paths = await download_images_v2([url], download_folder, session)
#
#                 if not pic_paths:
#                     need_clear = True
#             # print(pic_paths)
#
#             # Отримання імені файлу з URL
#             parsed_url = urlparse(url)
#             file_name = os.path.basename(unquote(parsed_url.path))
#             renamed = False
#
#             if not need_clear:
#                 name, file_format = file_name.rsplit(".", 1)
#                 while True:
#                     new_file_name = gen_rand_text() + "." + file_format
#                     if new_file_name in claimed_names:
#                         continue
#                     claimed_names.append(new_file_name)
#                     if not pic_paths:
#                         print('pic_paths', pic_paths)
#                         break
#                     if not pic_paths[0]:
#                         print('pic_paths[0]', pic_paths)
#                         break
#                     if pic_paths[0] == []:
#                         print('pic_paths[0]==[]', pic_paths)
#                         break
#                     if isinstance(pic_paths[0], str):
#                         os.rename(pic_paths[0], download_folder + "/" + new_file_name)
#                     if isinstance(pic_paths[0], list):
#                         os.rename(pic_paths[0][0], download_folder + "/" + new_file_name)
#                     break
#
#             # Process exactly image tags
#             for html_img_tag in soup.find_all("img"):
#                 try:
#                     html_img_url = html_img_tag["src"]
#                     if not html_img_url:
#                         raise ValueError()
#                 except:
#                     # Remove img tag if no src
#                     print(f"No src in img tag {html_img_tag}")
#                     if html_img_tag.parent.name == "p":
#                         try:
#                             html_img_tag.parent.replace_with("")
#                         except ValueError:
#                             newtext = re.sub(unescape(str(html_img_tag.parent)), '', unescape(str(soup)))
#                             soup = BeautifulSoup(newtext, "html.parser")
#                     else:
#                         try:
#                             html_img_tag.replace_with("")
#                         except ValueError:
#                             newtext = re.sub(unescape(str(html_img_tag)), '', unescape(str(soup)))
#                             soup = BeautifulSoup(newtext, "html.parser")
#
#                     continue
#                 # print(html_img_tag, html_img_url, unquote(html_img_url))
#
#                 if need_clear:
#                     # Find bad image
#                     if file_name in unquote(html_img_url):
#                         # Remove link from html
#                         if html_img_tag.parent.name == "p":
#                             # print("p", html_img_tag.parent, html_img_tag.parent.parent)
#                             try:
#                                 html_img_tag.parent.replace_with("")
#                             except ValueError:
#                                 # print(html_img_tag.parent, html_img_tag.parent.parent)
#                                 newtext = re.sub(unescape(str(html_img_tag.parent)), '', unescape(str(soup)))
#                                 soup = BeautifulSoup(newtext, "html.parser")
#                         else:
#                             # print("not p", html_img_tag)
#                             try:
#                                 html_img_tag.replace_with("")
#                             except ValueError:
#                                 newtext = re.sub(unescape(str(html_img_tag)), '', unescape(str(soup)))
#                                 soup = BeautifulSoup(newtext, "html.parser")
#                         # print(parsed_url, file_name, html_img_url)
#                     continue
#
#                 # Anyway rename
#                 if file_name in unquote(html_img_url):
#                     link_start = unquote(html_img_url).split(file_name)[0]
#                     link_new = link_start + new_file_name
#                     html_img_tag["src"] = link_new
#
#                     nl = html_img_tag["src"]
#                     new_links.append(nl)
#                     # print(f"fn {file_name}")
#                     # print(f"start {link_start}")
#                     # print(f"nfn {new_file_name}")
#                     # print(f"old src {html_img_url}")
#                     # print(f"new src {nl}")
#                     renamed = True
#
#             # Clear other tags from bad url
#             for html_img_tag in soup.find_all():
#                 if file_name in unquote(str(html_img_tag)):
#                     new_tag = unquote(str(html_img_tag))
#                     new_tag = new_tag[:new_tag.rfind("\"", 0, new_tag.find(file_name)) + 1] + new_tag[new_tag.find("\"", new_tag.find(file_name)):]
#                     try:
#                         html_img_tag.replace_with(new_tag)
#                     except ValueError:
#                         newtext = re.sub(unescape(str(html_img_tag)), new_tag, unescape(str(soup)))
#                         soup = BeautifulSoup(newtext, "html.parser")
#
#             if not need_clear and not renamed:
#                 print(f"Can not replace photo in excel (bo nema tega src), additional info: file_name {file_name}, new_file_name {new_file_name}, pic_paths {pic_paths}, url {url}")
#                 try:
#                     os.remove(download_folder + "/" + new_file_name)
#                     print(f"Photo file deleted")
#                 except:
#                     print(f"[Error] Can not delete photo file on path {download_folder}/{new_file_name}")
#
#             # print(len(list(soup.find_all("img"))))
#
#             sheet.cell(i_row, content_index, unescape(str(soup)))
#
#         if new_links:
#             sheet.cell(i_row, image_now_urls_index, " ".join(new_links))
#
#         i_row += 1
#
#
#     # ЗВУК ПРИ КІНЦІ ПАРСИНГУ
#     # playsound(r'./audio/rozpochala_parsing.mp3')
#     print("Downloading ended")
# # Скачування/завантаження картинок. Кінец функції завантаження.
#
#     # zip_file_path = f'static/results/archive_{random.randint(1000, 9999)}'
#     # zip_filename = shutil.make_archive(zip_file_path, 'zip', 'static/v2')
#
#     xlsx_file_path = f'{download_folder}/table_{random.randint(1000, 9999)}.xlsx'
#     uploaded_workbook.save(xlsx_file_path)
#
#     end_zip_file_path = f'static/archive_{random.randint(1000, 9999)}'
#     zip_filename = shutil.make_archive(end_zip_file_path, 'zip', download_folder)
#
#     return FileResponse(
#         path=end_zip_file_path + '.zip',
#         media_type='application/zip',
#         filename=os.path.basename(end_zip_file_path),
#         headers={
#             'Content-Disposition': f'attachment; filename="{os.path.basename(end_zip_file_path)}"'
#         }
#     )
# Скачування/завантаження картинок. Кінець коду.

@app.post("/crop_images")
async def crop_images(request: Request):
    data = await request.form()
    print(data)

    num_of_pixels = int(data.get('num_of_pixels', 100))
    is_left_crop = data.get('is_left_crop') == 'true'
    is_top_crop = data.get('is_top_crop') == 'true'
    is_right_crop = data.get('is_right_crop') == 'true'
    is_down_crop = data.get('is_down_crop') == 'true'
    need_mirroring = data.get('need_mirroring') == 'true'
    need_deleting = data.get('need_deleting') == 'true'
    file = data.get('file')

    folder_name = data.get('folder_name', 'images_for_cropping')

    if not Path(folder_name).exists():
        print(f'Error: nema takoj papki')
        return HTMLResponse(content="<h1>Error: No such folder</h1>", status_code=400)

    if need_deleting:
        if not file:
            print(f'Error: ne vkazali file')
            return HTMLResponse(content="<h1>Error: Ne vkazali file</h1>", status_code=400)

        print("Process started")

        elems_for_del = []
        for elem in Path(folder_name).iterdir():
            elems_for_del.append(elem.name)

        contents = await file.read()
        uploaded_workbook = load_workbook(filename=BytesIO(contents))
        sheet = uploaded_workbook._sheets[0]
        headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки

        content_index = headers.index("Content") + 1
        orig_urls_index = headers.index("Image Url_original") + 1

        i_row = 2
        cnt = 0

        import requests
        session = requests.session()

        while i_row <= sheet.max_row:
            content = sheet.cell(i_row, content_index).value  # Беремо тільки "Content"
            # orig_urls = sheet.cell(i_row, orig_urls_index).value  # Беремо тільки "Image Url_original"

            if cnt >= 1000:
                break

            if not content:
                cnt += 1
                continue

            soup = BeautifulSoup(content, "html.parser")
            # new_orig_urls: list = orig_urls.split(" ")
            for file_name in elems_for_del:
                # print(url)
                # parsed_url = urlparse(url)
                # file_name = os.path.basename(unquote(parsed_url.path))
                # if file_name in elems_for_del:
                #     new_orig_urls.remove(url)
                    for html_img_tag in soup.find_all("img"):
                        try:
                            html_img_url = html_img_tag["src"]
                        except:
                            print(f"No src in tag {html_img_tag}")
                            continue
                        # print(html_img_tag, html_img_url, unquote(html_img_url))

                        # Find bad image
                        if file_name in unquote(html_img_url):
                            # Remove link from html
                            if html_img_tag.parent.name == "p":
                                # print("p")
                                html_img_tag.parent.replace_with("")
                            else:
                                # print("not p")
                                html_img_tag.replace_with("")
                            # print(parsed_url, file_name, html_img_url)
                            # input()

                    # print(len(list(soup.find_all("img"))))

                    sheet.cell(i_row, content_index, str(soup))
                    # sheet.cell(i_row, orig_urls_index, " ".join(new_orig_urls))

            i_row += 1


        print("Process ended")

        xlsx_file_path = f'static/table_{random.randint(1000, 9999)}.xlsx'
        uploaded_workbook.save(xlsx_file_path)

        return FileResponse(
            path=xlsx_file_path,
            media_type='application/xlsx',
            filename=os.path.basename(xlsx_file_path),
            headers={
                'Content-Disposition': f'attachment; filename="{os.path.basename(xlsx_file_path)}"'
            }
        )


    for elem in Path(folder_name).iterdir():
        try:
            img = Image.open(elem.absolute())
        except UnidentifiedImageError:
            print(f'Error: can not open file {elem.absolute()}')
            continue

        print(f'Obrabotka file {elem}, width: {img.width},  height: {img.height}')

        try:
            img = img.crop((
                    0+(int(img.width*num_of_pixels/100)*is_left_crop),
                    0+(int(img.height*num_of_pixels/100)*is_top_crop),
                    img.width-(int(img.width*num_of_pixels/100)*is_right_crop),
                    img.height-(int(img.height*num_of_pixels/100)*is_down_crop)
                )
            )
        except ValueError:
            # якщо наприклад в картинці 100 пікселів, а ми хочемо відрізати 200
            print(f'Can not crop {elem} because too much pixels to cut')

        if need_mirroring:
            img = img.transpose(Image.FLIP_LEFT_RIGHT)

        print(f'Nova shirina: {img.width}, nova visota: {img.height}')

        try:
            img.save(elem.absolute())
        except Exception as e:
            print(f'Ne zmoglu sohranit file {elem}, error type: {type(e)}, error text: {e}')

    return HTMLResponse(status_code=200)


@app.post("/replace_text")
async def replace_text(request: Request):
    data = await request.form()
    print(data)

    prev_text = data.get('prev_text', None)
    next_text = data.get('next_text', None)

    if not prev_text or not next_text:
        print(f'Error: vvedit text v oba pola')
        return HTMLResponse(content="<h1>Error: vvedit text v oba pola</h1>", status_code=400)

    file = data.get('file')
    if not file:
        print(f'Error: ne vkazali file')
        return HTMLResponse(content="<h1>Error: Ne vkazali file</h1>", status_code=400)

    print("Process started")

    contents = await file.read()
    uploaded_workbook = load_workbook(filename=BytesIO(contents))
    data_sheet = uploaded_workbook._sheets[0]

    headers = [cell.value for cell in data_sheet[1]]  # Перший рядок містить заголовки

    content_index = headers.index("Content") # starts from 0

    for row_i, row in enumerate(data_sheet.iter_rows()):
        if row_i == 0: # skip headers
            continue

        for cell_i, cell in enumerate(row):
            c_v = str(cell.value).strip()
            data_sheet.cell(row_i+1, cell_i+1, str(c_v.replace(prev_text, next_text)))

            if cell_i == content_index:
                content = data_sheet.cell(row_i+1, cell_i+1).value
                content = str(content).strip()
                soup = BeautifulSoup(content, "html.parser")

                for html_img_tag in soup.find_all("img"):
                    if html_img_tag.parent.name == "p":
                        if "left" in str(html_img_tag.parent) or "right" in str(html_img_tag.parent) or "center" in str(html_img_tag.parent):
                            continue
                    else:
                        if "left" in str(html_img_tag) or "right" in str(html_img_tag) or "center" in str(html_img_tag):
                            continue

                    html_img_tag["class"] = "aligncenter"

                    data_sheet.cell(row_i+1, cell_i+1, str(soup))

    print("Process ended")

    xlsx_file_path = f'static/table_{random.randint(1000, 9999)}.xlsx'
    uploaded_workbook.save(xlsx_file_path)

    return FileResponse(
        path=xlsx_file_path,
        media_type='application/xlsx',
        filename=os.path.basename(xlsx_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(xlsx_file_path)}"'
        }
    )


@app.get("/adoptTitles", response_class=HTMLResponse)
async def adoptTitles(request: Request):
    return period_check(templates.TemplateResponse("adoptTitles.html", {"request": request}))


# temp name
@app.post("/process_excel_3")
async def process_excel_3(request: Request):
    data = await request.form()
    print(data)

    num_results = int(data.get('num_results'
                               '', 10))
    file = data.get('file')
    is_alt = data.get('isAlt') == 'true'
    from_save = data.get('fromSave') == 'true'

    if from_save:
        try:
            uploaded_workbook = load_workbook(filename="static/last_excel_save.xlsx")
        except:
            print("Can not load excel save file")
            return HTMLResponse(content="Error: Can not load excel save file", status_code=400)
    else:
        if file == "undefined":
            print("Need file for this operation")
            return HTMLResponse(content="Error: Need file for this operation", status_code=400)

        try:
            # Завантажуємо файл Excel
            contents = await file.read()
            contents = BytesIO(contents)

            # Спробуємо відкрити файл
            uploaded_workbook = load_workbook(filename=contents)
        except:
            print("Can not load excel file")
            return HTMLResponse(content="Error: Can not load excel file", status_code=400)

    # ЗВУК ПРИ ПОЧАТКУ ПАРСИНГУ
    playsound(r'./audio/rozpochala_parsing.mp3')

    try:
        data_sheet = uploaded_workbook._sheets[0]
        wrong_titles_sheet = uploaded_workbook._sheets[1]
        keywords_sheet = uploaded_workbook._sheets[2]
        blacklist_titles_sheet = uploaded_workbook._sheets[3]
        handle_titles_sheet = uploaded_workbook._sheets[4]
    except IndexError:
        print("Need at least 5 sheets")
        return HTMLResponse(content="Error: Need at least 5 sheets", status_code=400)

    headers = [cell.value for cell in data_sheet[1]]  # Перший рядок містить заголовки
    title_index = headers.index("Title") + 1
    content_index = headers.index("Content") + 1
    url_index = headers.index("URL") + 1

    is_no_keywords = False
    if str(keywords_sheet.cell(1, 1).value) == "STOP":
        is_no_keywords = True
    elif not str(keywords_sheet.cell(1, 1).value):
        print("No data in third sheet")
        return HTMLResponse(content="Error: No data in third sheet", status_code=400)

# Видалення строк з екселя DELETE. Розбити файл ексель по доменах на декілька файлів. Початок функції.
    if not is_alt:
        wrong_titles = []
        i_row = 2
        cnt = 0
        while i_row <= wrong_titles_sheet.max_row:
            wrong_title = wrong_titles_sheet.cell(i_row, 1).value
            i_row += 1

            if cnt >= 1000:
                break

            if wrong_title is None:
                cnt += 1
                continue
            #Функція переобходу заголовків з вкладки: Переобход заголовків
            wrong_title = str(wrong_title).strip()
            if wrong_title:
                wrong_titles.append(wrong_title)

            # Фукція заміни не правильних заголовків. Закоментована.
            #wrong_title = str(wrong_title).strip()
            #re_check = wrong_titles_sheet.cell(i_row - 1, 3).value
            #if str(re_check).strip() == "Zamina":
                #new_title = str(wrong_titles_sheet.cell(i_row - 1, 2).value).strip()
                #print("new", new_title)
                #re_wrong_titles_map[new_title] = wrong_title
                #wrong_titles.append(wrong_title)

            re_check = wrong_titles_sheet.cell(i_row - 1, 3).value
            if str(re_check).strip() == "delete":
                new_title = str(wrong_titles_sheet.cell(i_row - 1, 2).value).strip()
                print("Process title for deleting:", new_title)
                wrong_titles.append(new_title)

        # print(wrong_titles)

        i_row = 2
        cnt = 0
        domains = set()
        while True:
            end = False

            while i_row <= data_sheet.max_row:
                title = data_sheet.cell(i_row, title_index).value
                url = data_sheet.cell(i_row, url_index).value

                i_row += 1

                if cnt >= 1000:
                    end = True
                    break

                if title is None:
                    cnt += 1
                    continue

                # print(i_row-1, title.strip())
                if title.strip() in wrong_titles:
                    data_sheet.delete_rows(i_row-1, 1)
                    break

                domain = url.strip().split("://", 1)[1].split("/")[0]
                domains.add(domain)

            if end:
                break

        uploaded_workbook.close()

        total_xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
        uploaded_workbook.save(total_xlsx_file_path)

        uploaded_workbook = load_workbook(filename=total_xlsx_file_path)
        data_sheet = uploaded_workbook._sheets[0]

        domains = sorted(list(domains))
        print(domains)
        # input()

        files = []
        for idx, domain in enumerate(domains):
            print(f"Processing domain {domain} ({idx+1}/{len(domains)})")
            part_xlsx_file_path = f'static/output_{domain}_{random.randint(1000, 9999)}.xlsx'

            with open(total_xlsx_file_path, "rb") as f:
                contents = f.read()
                contents = BytesIO(contents)

                part_workbook = load_workbook(filename=contents)

                data_ws = part_workbook._sheets[0]
                data_ws.delete_rows(2, 2000)

                new_i_row = 2
                old_i_row = 2
                cnt = 0
                while old_i_row <= data_sheet.max_row:
                    url = data_sheet.cell(old_i_row, url_index).value

                    old_i_row += 1

                    if cnt >= 1000:
                        break

                    if url is None:
                        cnt += 1
                        continue

                    row_url = str(url).strip()

                    if domain in row_url:
                        for ic, cell in enumerate(data_sheet[old_i_row-1]):
                            data_ws.cell(new_i_row, ic+1, cell.value)
                        new_i_row += 1

                part_workbook.close()
                part_workbook.save(part_xlsx_file_path)

            files.append(part_xlsx_file_path)

        uploaded_workbook.close()

        zip_file_path = f'static/archive_{random.randint(1000, 9999)}.zip'
        create_zip_archive(files, zip_file_path)

        return FileResponse(
            path=zip_file_path,
            media_type='application/zip',
            filename=os.path.basename(zip_file_path),
            headers={
                'Content-Disposition': f'attachment; filename="{os.path.basename(zip_file_path)}"'
            }
        )


#Видалення строк з екселя. Розбити файл ексель по доменах на декілька файлів. Кінець функції.

    keywords = []
    wrong_titles = []
    blacklist_titles = []
    changed_titles = []
    handle_titles = []
    data_sheet_start_row_id = 2

    try:
        if from_save:
            try:
                with open("static/last_excel_save.json", "r", encoding="utf-8") as f:
                    loaded_save = json.load(f)
                print("Loaded data from json save file")
            except:
                print("Can not load json save file")
                return HTMLResponse(content="Error: Can not load json save file", status_code=400)

            keywords = loaded_save.get("keywords", [])
            wrong_titles = loaded_save.get("wrong_titles", [])
            blacklist_titles = loaded_save.get("blacklist_titles", [])
            changed_titles = loaded_save.get("changed_titles", [])
            handle_titles = loaded_save.get("handle_titles", [])
            data_sheet_start_row_id = loaded_save.get("data_sheet_start_row_id", 2)

        else:
            keywords = []
            i_row = 1
            cnt = 0
            while i_row <= keywords_sheet.max_row:
                word = keywords_sheet.cell(i_row, 1).value
                i_row += 1

                if cnt >= 1000:
                    break

                if word is None:
                    cnt += 1
                    continue

                word = str(word).strip()
                if word:
                    keywords.append(word + " ")

            wrong_titles = []
            i_row = 2
            cnt = 0
            while i_row <= wrong_titles_sheet.max_row:
                wrong_title = wrong_titles_sheet.cell(i_row, 1).value
                i_row += 1

                if cnt >= 1000:
                    break

                if wrong_title is None:
                    cnt += 1
                    continue

                # Функція Переобход заголовків повторно
                wrong_title = str(wrong_title).strip()
                if wrong_title:
                    wrong_titles.append(wrong_title)


            blacklist_titles = []
            i_row = 1
            cnt = 0
            while i_row <= blacklist_titles_sheet.max_row:
                blacklist_title = blacklist_titles_sheet.cell(i_row, 1).value
                i_row += 1

                if cnt >= 1000:
                    break

                if blacklist_title is None:
                    cnt += 1
                    continue

                blacklist_title = str(blacklist_title).strip()
                if blacklist_title:
                    blacklist_titles.append(blacklist_title)

            changed_titles = dict()
            handle_titles = []


        i_row = data_sheet_start_row_id
        cnt = 0
        print(wrong_titles)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options())#, desired_capabilities=capabilities)
        while i_row <= data_sheet.max_row:
            title = data_sheet.cell(i_row, title_index).value
            content = data_sheet.cell(i_row, content_index).value

            data_sheet_start_row_id = i_row

            if cnt >= 1000:
                break

            if title is None:
                cnt += 1
                continue

            title = str(title).strip()
            # if title in re_wrong_titles_map:
            #     print(f"переобход {re_wrong_titles_map[title]}")
            #     title = re_wrong_titles_map[title]
            #     data_sheet.cell(i_row, title_index, value=title)

            if (title in wrong_titles) or not wrong_titles:
                print(f"[row {i_row}] Working with wrong title '{title}'")
                target_kw = None
                if not is_no_keywords:
                    for kw in keywords:
                        if kw.lower() in title.lower():
                            target_kw = kw
                            break
                    # else:
                    #     handle_titles.append(title)
                    #     i_row +=1
                    #     continue

                # if is_alt:
                search_results = get_google_search_results_alt(title, num_results, need_titles=True, driver=driver)#, proxy_chrome=proxy_chrome)
                # else:
                #     search_results = get_google_search_results(title, need_titles=True)
                print(f"Get this search results: {json.dumps(search_results, ensure_ascii=False, indent=4)}")

                filtrated = []
                for x in search_results:
                    if is_no_keywords:
                        filtrated.append([None, x])
                    elif not target_kw:
                        # опрацювати фільтер по всім словам якщо немає слова з фільтрів у заголовку
                        for kw in keywords:
                            if kw.lower() in x.lower():
                                filtrated.append([target_kw, x])
                                break
                    elif target_kw.lower() in x.lower():
                        filtrated.append([target_kw, x])
                        print(f"Find this keyword: '{target_kw}' for search title '{x}'")
                        # break

                if not filtrated:
                    print(f"[Skipping] when parsing new titles for old title '{title}': no one search result have filter words")
                    handle_titles.append(title)
                    i_row += 1
                    continue

                alph = (
                    "abcdefghijklmnopqrstuvwxyz"
                    "1234567890"
                    "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"
                    "ґєії' "
                )
                cleaned_titles = []
                max_r = 0
                max_t = ""
                for kw, x in filtrated:
                    if kw:
                        x = x[x.lower().find(kw):]
                    x.replace("-", " ").replace("+", " ")
                    cleaned = ""
                    for l in x:
                        if l.lower() not in alph:
                            break
                        cleaned += l

                    cleaned = cleaned.strip()

                    print(f"Clean search title '{x}' to '{cleaned}'")

                    if cleaned in blacklist_titles:
                        continue

                    if len(cleaned.split(" ")) < 3:
                        continue

                    cleaned_titles.append(cleaned)

                    soup = BeautifulSoup(content, "html.parser")
                    for t in soup.find_all("h1"): # вирізаємо h1 з заголовком
                        t.replace_with("")

                    content = unescape(str(soup))

                    ratio = fuzz.ratio(title, cleaned)
                    print(f"Title ratio = {ratio}")

                    # если старый и новый заголовок почти одинаковы, то берём его сразу, не смотря в статью
                    # по принципу "хотя бы не сделать хуже"
                    if ratio >= 90:
                        max_r = ratio
                        max_t = cleaned
                        continue

                    skip = False
                    # проверяем каждое слово из гугловского заголовка на факт наличия в статье (даже если другое окончание и тд)
                    # если не нашли ни одного совпадения хотя бы на 80, то значит такого слова нет в тексте
                    for word1 in cleaned.split(" "):
                        found = False
                        for word2 in cleaned.split(" "):
                            if fuzz.ratio(word1, word2) >= 80:
                                found = True
                                break
                        if not found:
                            skip = True
                            break

                    if skip:
                        print(f"Пропускаем заголовок, потому что слова {word1} нет в тексте статьи")
                        continue

                    ratio = fuzz.partial_ratio(content, cleaned)
                    print(f"Content ratio = {ratio}")
                    if ratio > max_r: # compare with content, not old title
                        max_r = ratio
                        max_t = cleaned

                if not max_t:
                    print(f"[Skipping] when filtering and etc no one search result have been chosen")
                    handle_titles.append(title)
                    i_row += 1
                    continue

                new_title = max_t
                new_title = new_title[0].upper() + new_title[1:]

                print(f"Setting new title from '{title}' to '{new_title}' (ratio is {max_r})")

                changed_titles[title] = new_title

                changed_i_row = 2
                changed_cnt = 0
                if changed_titles:
                    if wrong_titles:
                        while changed_i_row <= wrong_titles_sheet.max_row:
                            c_wrong_title = wrong_titles_sheet.cell(changed_i_row, 1).value

                            if changed_cnt >= 1000:
                                break

                            if c_wrong_title is None:
                                changed_cnt += 1
                                continue

                            c_wrong_title = str(c_wrong_title).strip()

                            if changed_titles.get(c_wrong_title):
                                # print(f"Старий заголовок: '{c_wrong_title}', новий: '{changed_titles.get(c_wrong_title)}'")
                                if c_wrong_title != str(changed_titles.get(c_wrong_title)).strip():
                                    wrong_titles_sheet.cell(changed_i_row, 2, str(changed_titles.get(c_wrong_title)))

                            changed_i_row += 1
                    else:
                        c_wrong_titles = list(changed_titles.keys())
                        for ci, c_wrong_title in enumerate(c_wrong_titles):
                            c_wrong_title = str(c_wrong_title).strip()

                            # wrong_titles_sheet.cell(ci + 2, 1, c_wrong_title)

                            if changed_titles.get(c_wrong_title):
                                # print(f"Старий заголовок: '{c_wrong_title}', новий: '{changed_titles.get(c_wrong_title)}'")
                                if c_wrong_title != str(changed_titles.get(c_wrong_title)).strip():
                                    wrong_titles_sheet.cell(ci + 2, 1, c_wrong_title)
                                    wrong_titles_sheet.cell(ci + 2, 2, str(changed_titles.get(c_wrong_title)).strip())

                data_sheet.cell(i_row, title_index, new_title.strip())

                if not data_sheet.cell(i_row, content_index).value:
                    print(f"ERROR NOT \"CONTENT\" DATA IN {i_row}:{content_index}")
                    i_row += 1
                    continue

                soup = BeautifulSoup(data_sheet.cell(i_row, content_index).value.strip(), "html.parser")

                c_0 = soup.contents[0]
                if isinstance(c_0, Tag) and c_0.name == "h1":
                    c_0.replace_with(f"<h1>{new_title}</h1>")
                    data_sheet.cell(i_row, content_index, unescape(str(soup)).strip())
                else:
                    data_sheet.cell(i_row, content_index, f"<h1>{new_title}</h1> " + unescape(str(soup)).strip())

                # робить копію (типу збереження) кожні 1 проходи/запити в гугл/заголовки
                if i_row % 1 == 0:
                    print(f"Заголовок збереженно в тимчасовий файл.. (row id: {i_row})")

                    dumped_save = dict()
                    dumped_save["keywords"] = keywords
                    dumped_save["wrong_titles"] = wrong_titles
                    dumped_save["blacklist_titles"] = blacklist_titles
                    dumped_save["changed_titles"] = changed_titles
                    dumped_save["handle_titles"] = handle_titles
                    dumped_save["data_sheet_start_row_id"] = data_sheet_start_row_id

                    with open("static/last_excel_save.json", "w", encoding="utf-8") as f:
                        json.dump(dumped_save, f, ensure_ascii=False, indent=4)

                    # uploaded_workbook.close()

                    xlsx_file_path = f'static/last_excel_save.xlsx'
                    uploaded_workbook.save(xlsx_file_path)

                time.sleep(random.uniform(2, 5)) # пауза між запитами в гугл пошуку

                if i_row % 10 == 0:
                    # пауза кожні 10 запитів в гугл пошук
                    rnd = random.uniform(30, 60)
                    print(f"Рандомна пауза кожних 10 запитів 30-60 сек (вибрало {rnd} секунд)")
                    time.sleep(rnd)

            i_row += 1

        try:
            driver.close()
            driver.quit()
        except:
            pass

        i_row = 2
        cnt = 0
        if changed_titles:
            if wrong_titles:
                while i_row <= wrong_titles_sheet.max_row:
                    wrong_title = wrong_titles_sheet.cell(i_row, 1).value

                    if cnt >= 1000:
                        break

                    if wrong_title is None:
                        cnt += 1
                        continue

                    wrong_title = str(wrong_title).strip()

                    if changed_titles.get(wrong_title):
                        # print(f"Старий заголовок: '{wrong_title}', новий: '{changed_titles.get(wrong_title)}'")
                        if wrong_title != str(changed_titles.get(wrong_title)).strip():
                            wrong_titles_sheet.cell(i_row, 2, str(changed_titles.get(wrong_title)))

                    i_row += 1
            else:
                wrong_titles = list(changed_titles.keys())
                for i, wrong_title in enumerate(wrong_titles):
                    wrong_title = str(wrong_title).strip()

                    # wrong_titles_sheet.cell(i+2, 1, wrong_title)

                    if changed_titles.get(wrong_title):
                        # print(f"Старий заголовок: '{wrong_title}', новий: '{changed_titles.get(wrong_title)}'")
                        if wrong_title != str(changed_titles.get(wrong_title)).strip():
                            wrong_titles_sheet.cell(i + 2, 1, wrong_title)
                            wrong_titles_sheet.cell(i+2, 2, str(changed_titles.get(wrong_title)).strip())
        # else:
        #     while i_row <= wrong_titles_sheet.max_row:
        #         try:
        #             list(changed_titles.keys())[i_row - 2]
        #         except:
        #             break
        #         wrong_title = wrong_titles_sheet.cell(i_row, 1, list(changed_titles.keys())[i_row-2])
        #
        #         if cnt >= 1000:
        #             break
        #
        #         if wrong_title is None:
        #             cnt += 1
        #             continue
        #
        #         wrong_title = str(wrong_title).strip()
        #
        #         if changed_titles.get(wrong_title):
        #             wrong_titles_sheet.cell(i_row, 2, str(changed_titles.get(wrong_title)))
        #
        #         i_row += 1

        for i, t in enumerate(handle_titles):
            handle_titles_sheet.cell(i+1, 1, t.strip())

    except:
        print("Get error when parsing titles: " + "".join(traceback.format_exception(*sys.exc_info())))

        dumped_save = dict()
        dumped_save["keywords"] = keywords
        dumped_save["wrong_titles"] = wrong_titles
        dumped_save["blacklist_titles"] = blacklist_titles
        dumped_save["changed_titles"] = changed_titles
        dumped_save["handle_titles"] = handle_titles
        dumped_save["data_sheet_start_row_id"] = data_sheet_start_row_id

        with open("static/last_excel_save.json", "w", encoding="utf-8") as f:
            json.dump(dumped_save, f, ensure_ascii=False, indent=4)

        uploaded_workbook.close()

        xlsx_file_path = f'static/last_excel_save.xlsx'
        uploaded_workbook.save(xlsx_file_path)

        print("Successfully make save after operation interrupted by error")
        return HTMLResponse(content="Get error and save file", status_code=400)

    else:
        # ЗВУК ПРИ ЗАКІНЧЕННІ ПОШУКУ
        playsound(r'./audio/zakincheno_zbir_site_googl.mp3')
        uploaded_workbook.close()

        xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
        uploaded_workbook.save(xlsx_file_path)

        return FileResponse(
            path=xlsx_file_path,
            media_type='application/xlsx',
            filename=os.path.basename(xlsx_file_path),
            headers={
                'Content-Disposition': f'attachment; filename="{os.path.basename(xlsx_file_path)}"'
            }
        )


@app.get("/addKeys", response_class=HTMLResponse)
async def addKeys(request: Request):
    return period_check(templates.TemplateResponse("addKeys.html", {"request": request}))


@app.post("/process_excel_keys")
async def process_excel_keys(request: Request):
    data = await request.form()
    print(data)

    file = data.get('file')

    # Завантажуємо файл Excel
    contents = await file.read()
    contents = BytesIO(contents)

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=contents)

    try:
        data_sheet = uploaded_workbook._sheets[0]
    except IndexError:
        print("No data sheet finded")
        return HTMLResponse(content="<h1>Error: No data sheet finded</h1>", status_code=400)

    headers = [cell.value for cell in data_sheet[1]]  # Перший рядок містить заголовки
    title_index = headers.index("Title") + 1
    content_index = headers.index("Content") + 1

    i_row = 2
    cnt = 0
    while i_row <= data_sheet.max_row:
        title = data_sheet.cell(i_row, title_index).value
        content = data_sheet.cell(i_row, content_index).value

        if cnt >= 1000:
            break

        if title is None:
            cnt += 1
            continue

        title = str(title).strip()

        add_block = f'<span style="color: #fff"> {title}</span>'

        soup = BeautifulSoup(content, "html.parser")

        def has_punctuation(text):
            # Закоментувати, щоб не додавало заголовки в яких , кома і дві крапки. Стати на строку return False і ctrl + . Обов'язково англ розклдака клави.
            return False
            for x in ",:":
                if x in text:
                    return True
            return False

        # Додати h1 в статю. Додає тільки заголовки без коми і двокрапки.
        if not has_punctuation(title):
            c_0 = soup.contents[0]
            if isinstance(c_0, Tag) and c_0.name == "h1":
                c_0.replace_with(f"<h1>{title}</h1>")
            else:
                soup = BeautifulSoup(f"<h1>{title}</h1> " + unescape(str(soup)).strip(), "html.parser")

        p_tags = soup.find_all("p")

        p_len = len(p_tags)
        if p_len < 3:
            print(f"Error: In title '{title}' content has < 3 paragraphs")
            i_row += 1
            continue

        p_first: Tag = p_tags[0]
        center_index = int(p_len/2) if p_len%2 == 0 else int(p_len/2) +1
        p_center = p_tags[center_index]
        p_last = p_tags[-1]

        p_first.append(add_block)
        # print(p_first)
        p_center.append(add_block)
        p_last.append(add_block)

        img_tags = soup.find_all("img")

        for x in img_tags:
            x["alt"] = title
            if title[-1] == ".":
                x["title"] = title[:-1] + ": Merlin"
            else:
                x["title"] = title + ": Merlin"

        data_sheet.cell(i_row, content_index, unescape(str(soup)))

        i_row += 1

    uploaded_workbook.close()

    xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
    uploaded_workbook.save(xlsx_file_path)

    return FileResponse(
        path=xlsx_file_path,
        media_type='application/xlsx',
        filename=os.path.basename(xlsx_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(xlsx_file_path)}"'
        }
    )



@app.get("/perelinkovka", response_class=HTMLResponse)
async def perelinkovka(request: Request):
    return period_check(templates.TemplateResponse("perelinkovka.html", {"request": request}))


xml_linkovka_text = """
<?xml version="1.0" encoding="UTF-8"?>
<relations>
<!-- ПЕРЕЛІНКОВКА ГОЛОВНИХ З ГОЛОВНИМИ -->
{main_to_main}


<!-- ПЕРЕЛІНКОВКА ГОЛОВНИХ З НЕ ГОЛОВНИМИ -->
{main_to_other}


<!-- ПЕРЕЛІНКОВКА НЕ ГОЛОВНІ СТАТІ З НЕ ГОЛОВНИМИ СТАТЯМИ МІЖ СТАТЯМИ -->
{other_to_other}

</relations>
"""


@app.post("/process_excel_perelinkovka")
async def process_excel_perelinkovka(request: Request):
    data = await request.form()
    print(data)

    file = data.get('file')

    # Завантажуємо файл Excel
    contents = await file.read()
    contents = BytesIO(contents)

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=contents)

    try:
        sh_1 = uploaded_workbook._sheets[0]
        sh_2 = uploaded_workbook._sheets[1]
        sh_3 = uploaded_workbook._sheets[2]
        sh_4 = uploaded_workbook._sheets[3]
    except IndexError:
        print("Not enough sheets finded")
        return HTMLResponse(content="<h1>Error: Not enough sheets finded</h1>", status_code=400)

    main_ids = []
    other_ids = []

    i_row = 2
    cnt = 0
    while i_row <= sh_1.max_row:
        article_id_cell = sh_1.cell(i_row, 1)

        if cnt > 5000:
            break

        if article_id_cell.value is None:
            cnt += 1
            i_row += 1
            continue

        if article_id_cell.font.b:
            main_ids.append(article_id_cell.value)
        else:
            other_ids.append(article_id_cell.value)

        i_row += 1

    print(f"Головних: {len(main_ids)}")
    print(f"Не головних: {len(other_ids)}")

    main_ids_from = main_ids[:-1]
    main_ids_to = main_ids[1:]

    xml_main_to_main = []
    for i, ids in enumerate(zip(main_ids_from, main_ids_to)):
        sh_2.cell(i + 2, 2, ids[0])
        sh_2.cell(i + 2, 4, ids[1])

        sh_2.cell(i + 2, 7, f'<relation from="{ids[0]}" to="{ids[1]}" both="true"/>')
        xml_main_to_main.append(f'<relation from="{ids[0]}" to="{ids[1]}" both="true"/>')

    coefficient = len(other_ids) // len(main_ids)
    if len(other_ids) % len(main_ids) != 0:
        coefficient += 1

    print(f"Коефіцієнт головних до неголовних: {coefficient}")

    cnt_1 = 0
    cnt_2 = 0
    xml_main_to_other = []
    for i, x in enumerate(other_ids):
        sh_3.cell(i + 2, 2, main_ids[i % len(main_ids)])
        sh_3.cell(i + 2, 4, x)

        sh_3.cell(i + 2, 7, f'<relation from="{main_ids[i % len(main_ids)]}" to="{x}" both="true"/>')
        xml_main_to_other.append(f'<relation from="{main_ids[i % len(main_ids)]}" to="{x}" both="true"/>')

        cnt_1 += 1
        if cnt_1 == coefficient - 1:
            cnt_1 = 0
            cnt_2 += 1

    multiply_other_ids = other_ids * len(main_ids)
    rnd_multiply_other_ids = multiply_other_ids[:]
    if len(rnd_multiply_other_ids) > 1_000_000:
        print("Обрізаємо 'не головні до не головних' до 1 млн рядків, бо більше не можна")

    random.shuffle(rnd_multiply_other_ids)
    rnd_multiply_other_ids = rnd_multiply_other_ids[:1_000_000]
    for x in range(10):
        flag = 0
        for i in range(len(rnd_multiply_other_ids) - 1):
            if rnd_multiply_other_ids[i] == multiply_other_ids[i]:
                flag += 1
                rnd_multiply_other_ids[i], rnd_multiply_other_ids[i + 1] = rnd_multiply_other_ids[i + 1], \
                rnd_multiply_other_ids[i]
                # break

        if flag == 0:
            break

    xml_other_to_other = []
    for i in range(len(rnd_multiply_other_ids)):
        sh_4.cell(i + 2, 2, multiply_other_ids[i])
        sh_4.cell(i + 2, 4, rnd_multiply_other_ids[i])

        sh_4.cell(i + 2, 7, f'<relation from="{multiply_other_ids[i]}" to="{rnd_multiply_other_ids[i]}" both="false"/>')
        xml_other_to_other.append(f'<relation from="{multiply_other_ids[i]}" to="{rnd_multiply_other_ids[i]}" both="false"/>')

        if (i + 1) % 10_000 == 0:
            print(f"Не головні з не головними: {i + 1} рядків оброблено")

    uploaded_workbook.close()

    xlsx_file_path = f'static/perelinkovka_{datetime.now().date()}.xlsx'
    uploaded_workbook.save(xlsx_file_path)

    xml_file_path = f'static/perelinkovka_{datetime.now().date()}.xml'
    global xml_linkovka_text
    xml_text = xml_linkovka_text.format(main_to_main="\n".join(xml_main_to_main),
                                        main_to_other="\n".join(xml_main_to_other),
                                        other_to_other="\n".join(xml_other_to_other))
    with open(xml_file_path, "w", encoding="utf-8") as f:
        f.write(xml_text.strip())

    files = [xlsx_file_path, xml_file_path]
    zip_file_path = f'static/archive_{random.randint(1000, 9999)}.zip'
    create_zip_archive(files, zip_file_path)

    return FileResponse(
        path=zip_file_path,
        media_type='application/zip',
        filename=os.path.basename(zip_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(zip_file_path)}"'
        }
    )


@app.get("/wordExcel", response_class=HTMLResponse)
async def wordExcel(request: Request):
    return period_check(templates.TemplateResponse("wordExcel.html", {"request": request}))

# Початок функції. Фільтрування категорій
@app.post("/process_excel_4")
async def process_excel_4(request: Request):
    data = await request.form()
    print(data)

    file = data.get('file')
    moreThanOneCategory = data.get('moreThanOneCategory') == 'true'

    # Завантажуємо файл Excel
    contents = await file.read()
    contents = BytesIO(contents)

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=contents)

    try:
        data_sheet = uploaded_workbook._sheets[0]
    except IndexError:
        print("No data sheet finded")
        return HTMLResponse(content="<h1>Error: No data sheet finded</h1>", status_code=400)

    try:
        keywords_sheet = uploaded_workbook._sheets[2]
    except IndexError:
        print("No keys sheet finded")
        return HTMLResponse(content="<h1>Error: No keys sheet finded</h1>", status_code=400)

    # headers = [cell.value for cell in data_sheet[1]]  # Перший рядок містить заголовки
    # title_index = headers.index("Title") + 1
    # content_index = headers.index("Content") + 1

    all_titles = []
    all_urls = []
    all_titles_words = []
    all_urls_words = []
    for ix, row in enumerate(data_sheet.iter_rows()):
        if ix == 0:
            continue

        t_v = str(row[2].value).strip()
        u_v = str(row[4].value).strip()
        if t_v:
            all_titles.append(t_v)
            all_titles_words += re.findall(r"[a-zA-Zа-яА-ЯёЁіІїЇєЄ']+", t_v)
        if u_v:
            all_urls.append(u_v)
            all_urls_words += re.findall(r"[a-zA-Zа-яА-ЯёЁіІїЇєЄ']+", u_v)
# Додати слова які не потрібно рахувати в статистиці, в лозі програми.
    black_list_words = ["chego", "chej", "chem", "chey", "chja", "chto", "chya", "gde", "kak", "kakoj", "kakov", "kakoy", "kogda", "komu", "komy", "kotoryj", "kotoryy", "kuda", "otchego", "otkuda", "otkyda", "pochemu", "skolko", "zachem", "где", "зачем", "как", "каков", "какой", "когда", "кому", "который", "куда", "откида", "откуда", "отчего", "почему", "сколько", "чего", "чей", "чем", "что", "чья", "від чого", "відкиду", "де", "звідки", "коли", "кому", "кому", "куди", "навіщо", "скільки", "чий", "чий", "чим", "чия", "чия", "чия", "чого", "чому", "що", "як", "який", "який", "який", "який", "який", "який"]

    titles_stat = dict()
    for word in all_titles_words:
        titles_stat[word] = titles_stat.get(word, 0) + 1

    urls_stat = dict()
    for word in all_urls_words:
        urls_stat[word] = urls_stat.get(word, 0) + 1

    for x in sorted(titles_stat.items(), key=lambda k: -k[1]):
        if x[0].lower() in black_list_words or x[1] < 2 or len(x[0]) < 3:
            continue

        print(f"{x[0]} - {x[1]}")

    for x in sorted(urls_stat.items(), key=lambda k: -k[1]):
        if x[0].lower() in black_list_words or x[1] < 2 or len(x[0]) < 3:
            continue

        print(f"{x[0]} - {x[1]}")


    i_row = 1
    cnt = 0
    keywords = []
    while i_row <= keywords_sheet.max_row:
        keyword = keywords_sheet.cell(i_row, 1).value

        if cnt >= 1000:
            break

        if keyword is None:
            cnt += 1
            continue

        keyword = str(keyword).strip()

        keywords.append(keyword)

        i_row += 1

    print(f"{len(keywords)} keywords collected")

    continue_filter = False
    try:
        ws = uploaded_workbook._sheets[6]
        continue_filter = True
    except:
        uploaded_workbook.create_sheet("Відфільтровані категорії", 6)
        ws = uploaded_workbook._sheets[6]
        ws.column_dimensions["A"].width = data_sheet.column_dimensions["A"].width
        ws.column_dimensions["B"].width = data_sheet.column_dimensions["B"].width
        ws.column_dimensions["C"].width = data_sheet.column_dimensions["C"].width
        ws.column_dimensions["D"].width = data_sheet.column_dimensions["D"].width
        ws.column_dimensions["E"].width = data_sheet.column_dimensions["E"].width


    def copy_row(row_for_copy, to_sheet, to_row_index):
        for el_i, src_cell in enumerate(row_for_copy):
            to_sheet.cell(to_row_index, el_i+1, src_cell.value)

            dst_cell = to_sheet.cell(to_row_index, el_i + 1)

            dst_cell.font = copy(src_cell.font)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)

            dst_cell.number_format = src_cell.number_format

    to_row = 1
    if continue_filter:
        for ix, row in enumerate(ws.iter_rows()):
            cn = 0
            for cell in row:
                if not (cell.value.strip()):
                    print(cell.value, ix)
                    input()
                if cell.value is None:
                    cn += 1
            if cn == len(row):
                to_row = ix + 1
                print(123, to_row)
                break
        if to_row == 1:
            to_row = len([row for row in ws.iter_rows()]) + 1
        else:
            to_row = min(to_row, len([row for row in ws.iter_rows()]) + 1)

        print(12345, to_row, len([row for row in ws.iter_rows()]) + 1)

    print(f"Start copy process")
    max_row = 0
    max_cell = 0
    rows_for_cut = []
    for ix, row in enumerate(data_sheet.iter_rows()):
        max_row += 1
        tmp = 0
        for cell in row:
            tmp += 1
            # if cell is None:
            #     continue

        if ix == 0 and not continue_filter:
            copy_row(row, ws, to_row)
            to_row += 1
            continue

        t_v = str(row[2].value).strip()
        u_v = str(row[4].value).strip()
        for kw in keywords:
            if kw.lower() in t_v.lower() or kw.lower() in u_v.lower():
                # copy row
                # sheet_i, to_row_i = kw_info[kw]
                copy_row(row, ws, to_row)
                ws.cell(to_row, 6, kw)
                to_row += 1
                rows_for_cut.append(ix+1)
                # kw_info[kw] = [sheet_i, to_row_i+1]
                if not moreThanOneCategory:
                    break

        if tmp > max_cell:
            max_cell = tmp

    print(rows_for_cut)
    rows_for_cut = list(sorted(list(set(rows_for_cut))))
    print(rows_for_cut)
    cutted_num = 0
    for old_row_excel_id in rows_for_cut:
        data_sheet.delete_rows(old_row_excel_id - cutted_num, 1)
        cutted_num += 1

    print(f"End copy process, max row num: {max_row}, max cell num: {max_cell}")

    # return HTMLResponse(status_code=200)

    uploaded_workbook.close()

    xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
    uploaded_workbook.save(xlsx_file_path)

    return FileResponse(
        path=xlsx_file_path,
        media_type='application/xlsx',
        filename=os.path.basename(xlsx_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(xlsx_file_path)}"'
        }
    )

# Кінець функції. Фільтрування категорій



class UpdatedContent(BaseModel):
    updated_contents: List[dict]

@app.post("/save_excel")
async def save_excel(updated_contents: UpdatedContent):
    global uploaded_workbook
    if uploaded_workbook is None:
        return JSONResponse(status_code=400, content={"error": "No file uploaded."})
    workbook_copy_buf = BytesIO()
    uploaded_workbook.save(workbook_copy_buf)
    workbook_copy_buf.seek(0)
    workbook_copy = load_workbook(workbook_copy_buf)
    sheet = workbook_copy.active
    headers = [cell.value for cell in sheet[1]]
    content_index = headers.index("Content") + 1
    title_index = headers.index("Title") + 1

    """
    Пошук і заміна слів + символів з файлу Python_Poisk_Zamena.json
    Заміна відбувається при нажимані на кнопку зберегти в Excel на сторінці візуалізації.
    """
    replacements_count = 0
    with open('templates/static/Python_Poisk_Zamena.json', 'r', encoding='utf-8') as f:
        replace_rules = json.load(f)
        rules_count = len(replace_rules)
        print(f"\nЗавантажено правил: {rules_count}")

    def replace_text(text: str) -> str:
        nonlocal replacements_count
        if not text:
            return text

        def find_word(text, word):
            pattern = rf"(?<=^|[\s\p{{P}}\p{{S}}\d]){regex.escape(word)}(?=[\s\p{{P}}\p{{S}}\d]|$)"
            return regex.search(pattern, text)

        for rule in replace_rules:
            while True:
                match = find_word(text, rule['from'])
                if not match:
                    break
                text = text[:match.span()[0]] + rule['to'] + text[match.span()[1]:]
                replacements_count += 1
        return text

    for row in range(2, sheet.max_row + 1):
        title = sheet.cell(row=row, column=title_index).value
        if title:
            new_title = replace_text(title)
            sheet.cell(row=row, column=title_index, value=new_title)

    i_row = 2
    for content in updated_contents.updated_contents:
        if not content or not content["content"]:
            sheet.delete_rows(idx=i_row)
            continue
        current_content = replace_text(content["content"])
        soup = BeautifulSoup(current_content, "html.parser")
        for x in soup.find_all("img"):
            next = None
            for i in range(x.parent.contents.index(x), len(x.parent.contents)):
                if not str(x.parent.contents[i]).strip():
                    continue
                next = x.parent.contents[i]
            if isinstance(next, NavigableString):
                if x.parent.name == "p":
                    x.replace_with(unescape(str(x)) + "</p>")
                else:
                    x.replace_with("<p>" + unescape(str(x)) + "</p>")
        sheet.cell(i_row, content_index, value=unescape(str(soup)))
        i_row += 1

    print(f"Виконано замін: {replacements_count}")

    # Пошук міст росії. Початок функції
    with open('templates/static/Mista_rusion.txt', 'r', encoding='utf-8') as f:
        mista = f.readlines()
        for x in range(len(mista)):
            mista[x] = mista[x].strip()
        mista = list(filter(lambda el: el.strip(), mista))

    def find_city(text: str):
        nonlocal mista
        finded_cities = []

        for m in mista:
            # v1
            # if m.lower() in text.lower():
            #     finded_cities.append(m)
            #     print(m)

            # v2
            # ix = text.find(m)
            # if ix != -1:
            #     if not text[ix-1].isalnum() and not text[ix+len(m)+1].isalnum():
            #         finded_cities.append(m)
            #         print(m)

            # v3
            pattern = rf"(?<=^|[\s\p{{P}}\p{{S}}\d]){regex.escape(m.lower())}(?=[\s\p{{P}}\p{{S}}\d]|$)"
            if regex.search(pattern, text.lower()):
                finded_cities.append(m)
                # print(m) # better off for fast work


        return finded_cities

    sheet.column_dimensions["A"].width = 15
    sheet.cell(row=1, column=1, value="Міста росії")
    cell = sheet.cell(1, 1)
    ft = copy(cell.font)
    ft.b = True
    ft.sz = 14
    cell.font = ft
    for row in range(2, sheet.max_row + 1):
        content = sheet.cell(row=row, column=content_index).value
        if content:
            finded_cities = find_city(content)
            cell = sheet.cell(row, 1)
            ft = copy(cell.font)
            ft.b = True
            ft.sz = 14
            cell.font = ft

            if finded_cities:
                sheet.cell(row=row, column=1, value=", ".join(finded_cities))
            else:
                sheet.cell(row=row, column=1, value=" ")

            if row % 1 == 0:
                print(f"Пошук слів з списку. Строка {row} оброблена, найдено: ({','.join(finded_cities)})")
        else:
            print(f"Строка {row} пропущена, бо пуста статя")

    # Пошук міст росії. Кінець функції

    output = BytesIO()
    workbook_copy.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=updated_content.xlsx"}
    )


@app.post("/process_excel_zamina")
async def process_excel_zamina(file: UploadFile = File(...)):
    global uploaded_workbook
    # Завантажуємо файл Excel
    contents = await file.read()

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=BytesIO(contents))

    # Отримання даних з першого листа
    sheet = uploaded_workbook.active

    # Отримуємо заголовки стовпців
    headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки

    # Індекси потрібних стовпців
    content_index = headers.index("Content")+1
    title_index = headers.index("Title") + 1

    """
    Пошук і заміна слів + символів з файлу Python_Poisk_Zamena.json
    Заміна відбувається при нажимані на кнопку зберегти в Excel на сторінці візуалізації.
    """
    replacements_count = 0
    with open('templates/static/Python_Poisk_Zamena.json', 'r', encoding='utf-8') as f:
        replace_rules = json.load(f)
        rules_count = len(replace_rules)
        print(f"\nЗавантажено правил: {rules_count}")

    def replace_text(text: str) -> str:
        nonlocal replacements_count
        if not text:
            return text

        def find_word(text, word):
            pattern = rf"(?<=^|[\s\p{{P}}\p{{S}}\d]){regex.escape(word)}(?=[\s\p{{P}}\p{{S}}\d]|$)"
            return regex.search(pattern, text)

        for rule in replace_rules:
            while True:
                match = find_word(text, rule['from'])
                if not match:
                    break
                text = text[:match.span()[0]] + rule['to'] + text[match.span()[1]:]
                replacements_count += 1
        return text

    for row in range(2, sheet.max_row + 1):
        title = sheet.cell(row=row, column=title_index).value
        if title:
            new_title = replace_text(title)
            sheet.cell(row=row, column=title_index, value=new_title)

    for i_row in range(2, sheet.max_row + 1):
        content = sheet.cell(row=i_row, column=content_index).value
        current_content = replace_text(content)
        if not current_content:
            continue
        soup = BeautifulSoup(current_content, "html.parser")
        for x in soup.find_all("img"):
            next = None
            for i in range(x.parent.contents.index(x), len(x.parent.contents)):
                if not str(x.parent.contents[i]).strip():
                    continue
                next = x.parent.contents[i]
            if isinstance(next, NavigableString):
                if x.parent.name == "p":
                    x.replace_with(unescape(str(x)) + "</p>")
                else:
                    x.replace_with("<p>" + unescape(str(x)) + "</p>")
        sheet.cell(i_row, content_index, value=unescape(str(soup)))

    print(f"Виконано замін: {replacements_count}")

    # Пошук міст росії. Початок функції

    # with open('templates/static/Mista_rusion.txt', 'r', encoding='utf-8') as f:
    #     mista = f.readlines()
    #     for x in range(len(mista)):
    #         mista[x] = mista[x].strip()
    #     mista = list(filter(lambda el: el.strip(), mista))
    #
    # def find_city(text: str):
    #     nonlocal mista
    #     finded_cities = []
    #
    #     for m in mista:
    #         # v1
    #         # if m.lower() in text.lower():
    #         #     finded_cities.append(m)
    #         #     print(m)
    #
    #         # v2
    #         # ix = text.find(m)
    #         # if ix != -1:
    #         #     if not text[ix-1].isalnum() and not text[ix+len(m)+1].isalnum():
    #         #         finded_cities.append(m)
    #         #         print(m)
    #
    #         # v3
    #         pattern = rf"(?<=^|[\s\p{{P}}\p{{S}}]){regex.escape(m.lower())}(?=[\s\p{{P}}\p{{S}}]|$)"
    #         if regex.search(pattern, text.lower()):
    #             finded_cities.append(m)
    #             # print(m) # better off for fast work
    #
    #
    #     return finded_cities
    #
    # sheet.column_dimensions["A"].width = 15
    # sheet.cell(row=1, column=1, value="Міста росії")
    # cell = sheet.cell(1, 1)
    # ft = copy(cell.font)
    # ft.b = True
    # ft.sz = 14
    # cell.font = ft
    # for row in range(2, sheet.max_row + 1):
    #     content = sheet.cell(row=row, column=content_index).value
    #     if content:
    #         finded_cities = find_city(content)
    #         cell = sheet.cell(row, 1)
    #         ft = copy(cell.font)
    #         ft.b = True
    #         ft.sz = 14
    #         cell.font = ft
    #
    #         if finded_cities:
    #             sheet.cell(row=row, column=1, value=", ".join(finded_cities))
    #         else:
    #             sheet.cell(row=row, column=1, value=" ")
    #
    #         if row % 1 == 0:
    #             print(f"Пошук слів з списку. Строка {row} оброблена, найдено: ({','.join(finded_cities)})")
    #     else:
    #         print(f"Строка {row} пропущена, бо пуста статя")

    # Пошук міст росії. Кінець функції

    output = BytesIO()
    uploaded_workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=updated_content.xlsx"}
    )


@app.post("/split_excel_content")
async def split_excel_content(file: UploadFile = File(...)):
    global uploaded_workbook
    # Завантажуємо файл Excel
    contents = await file.read()

    # Спробуємо відкрити файл
    main_workbook = load_workbook(filename=BytesIO(contents))
    split_workbook = load_workbook(filename=BytesIO(contents))

    m_ws = main_workbook._sheets[0]
    s_ws = split_workbook._sheets[0]

    # Розділення статей на частини. Початок функції

    s_ws.delete_rows(2, 10_000)
    s_i_row = 2

    def split_html(html, max_chars=5000):
        soup = BeautifulSoup(html, "html.parser")

        chunks = []
        current_soup = BeautifulSoup("", "html.parser")
        current_text_length = 0

        for element in soup.contents:
            element_text = element.get_text(strip=True)

            if current_text_length + len(element_text) > max_chars:
                chunks.append(str(current_soup))
                current_soup = BeautifulSoup("", "html.parser")
                current_text_length = 0

            current_soup.append(copy(element))
            current_text_length += len(element_text)

        if current_text_length > 0:
            chunks.append(str(current_soup))

        return chunks

    for ix, row in enumerate(m_ws.iter_rows()):
        if ix == 0:
            continue
        id_ = row[1].value
        content = row[3].value
        if not id_ or not content:
            continue

        print(f"Стаття з ID {id_}: приблизно символів всього: {len(content)}, приблизно символів без розмітки: {len(html2text(content))}")

        content = content.replace("<html>", "").replace("<body>", "").replace("</html>", "").replace("</body>", "")

        # Тут 10000 це мінімальний розмір входящого блоку, щоб пустити його на розбивку. Замість 10000 поставити -1, щоб розбивало всі підряд
        if len(html2text(content)) < 10000:
            for iy, el in enumerate(row):
                s_ws.cell(s_i_row, iy + 1, el.value)
            s_i_row += 1
            continue

        # Тут 5000 це максимальний розмір блоків на виході чистим текстом без розмітки
        chunks = split_html(content, 10000)
        print(f"Статтю розбито на стільки блоків: {len(chunks)}")

        for i_ch, chunk in enumerate(chunks):
            init_row = [" "] * len(row)
            if i_ch == 0:
                for iy, el in enumerate(row):
                    init_row[iy] = el.value
            init_row[1] = f"{id_}.{i_ch + 1}"
            init_row[3] = unescape(chunk)

            for iy, el in enumerate(init_row):
                s_ws.cell(s_i_row, iy + 1, el)
            s_i_row += 1

    # Розділення статей на частини. Кінець функції

    output = BytesIO()
    split_workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=splited_content.xlsx"}
    )


@app.post("/concatenate_excel_content")
async def concatenate_excel_content(file: UploadFile = File(...)):
    global uploaded_workbook
    # Завантажуємо файл Excel
    contents = await file.read()

    # Спробуємо відкрити файл
    main_workbook = load_workbook(filename=BytesIO(contents))
    concatenate_workbook = load_workbook(filename=BytesIO(contents))

    m_ws = main_workbook._sheets[0]
    c_ws = concatenate_workbook._sheets[0]

    # Збірка статей з частин. Початок функції

    c_ws.delete_rows(2, 10_000)
    c_i_row = 2

    concat_row = None

    for ix, row in enumerate(m_ws.iter_rows()):
        if ix == 0:
            continue
        id_ = row[1].value
        content = row[3].value
        if not id_ or not content:
            continue

        print(f"Стаття з ID {id_}: приблизно символів всього: {len(content)}, приблизно символів без розмітки: {len(html2text(content))}")

        if len(id_.split(".")) == 2:
            if concat_row:
                for iy, el in enumerate(concat_row):
                    c_ws.cell(c_i_row, iy + 1, el)
                c_i_row += 1
                concat_row = None
            for iy, el in enumerate(row):
                c_ws.cell(c_i_row, iy + 1, el.value)
            c_i_row += 1
            continue

        if id_.split(".")[-1] == "1":
            if concat_row:
                for iy, el in enumerate(concat_row):
                    c_ws.cell(c_i_row, iy + 1, el)
                c_i_row += 1

            concat_row = [" "] * len(row)

            for iy, el in enumerate(row):
                concat_row[iy] = el.value

            concat_row[1] = ".".join(id_.split(".")[:-1])
        else:
            concat_row[3] += row[3].value

    if concat_row:
        for iy, el in enumerate(concat_row):
            c_ws.cell(c_i_row, iy + 1, el)
        c_i_row += 1

    # Збірка статей з частин. Кінець функції

    output = BytesIO()
    concatenate_workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=concatenated_content.xlsx"}
    )


# ДУМАЮ ТУТ МОЖЕТЕ ВСТАВИТИ СВІЙ КОД И КОД ПОПЕРЕДНЬОГО ВИКОНАВЦЯ. ТІЛЬКИ КОМЕНТ ТУТ ЗАЛИШТЕ
# МОЖЛИВО ЦЕ НЕ ПРАВИЛЬНЕ МІСЦЕ ВАМ ВИДНІШЕ. КНОПКА ЯКУ КОПІЮЄМО В adoptTitles.html

# Перевірка дублікатів заголовків


@app.post("/remove_duplicates")
async def remove_duplicates(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    num_results: int = Form(...),
):
    # file = data.get('file')

    if file == "undefined":
        print("Need file for this operation")
        return HTMLResponse(content="Error: Need file for this operation", status_code=400)

    try:
        # Завантажуємо файл Excel
        contents = await file.read()
        contents = BytesIO(contents)

        # Спробуємо відкрити файл
        uploaded_workbook = load_workbook(filename=contents)
    except:
        print("Can not load excel file")
        return HTMLResponse(content="Error: Can not load excel file", status_code=400)

    ws_1 = uploaded_workbook._sheets[0]


    # Видалення дублікатів через rapidfuzz. початок функції
    total_stati_count = 0
    for r in ws_1.iter_rows():
        if r[2].value is None:
            continue
        total_stati_count += 1

    removed_stati_count = 0
    while True:
        # break
        flag = False
        for ix, rowx in enumerate(ws_1.iter_rows()):
            if ix == 0 or rowx[2].value is None:
                continue
            titlex = str(rowx[2].value).strip()
            while True:
                flag = False
                for iy, rowy in enumerate(ws_1.iter_rows()):
                    if iy == 0 or rowy[2].value is None:
                        continue
                    titley = str(rowy[2].value).strip()
                    if ix == iy:
                        continue
                    #
                    #print(titlex, titley, (("," in titlex or ":" in titlex) or ("," in titley or ":" in titley)) and fuzz.ratio(str(titlex).strip(), str(titley).strip()))

                    # ТАК НЕ ПІДХОДИТЬ if fuzz.ratio(str(titlex).strip(), str(titley).strip()) >= 80:

                    if (
                            # перша і друга функція # Це удаляє дублі які мають рівно 95 або більше процентів схожесті.
                            # Працюють по черзі. Спочатку закоментувати третю. І or теж. Включити бота. Потім закоментувати 1 і 2 функцію. І розкоментувати третю функцію. Включити бота.
                            # Закоментувати спочатку першу і другу одну, включити бота, потім
                            fuzz.ratio(str(titlex).strip(), str(titley).strip()) >= 97 # БУЛО >= 98, 95
                            or
                            fuzz.ratio(" ".join(set(str(titlex).strip().split())), " ".join(set(str(titley).strip().split()))) >= 95 # Було >= 95, потім 40. Удаляє дублі в яких слова в заголовці міняються місьцями. Рівно 95 або більше процентів.
                            # третя
                            #or
                            #((("," in titlex or ":" in titlex) and ("," in titley or ":" in titley)) and fuzz.ratio(str(titlex).strip(), str(titley).strip()) < 40)
                    ):

                        flag = True
                        ws_1.delete_rows(iy + 1, 1)
                        removed_stati_count += 1

                        print(titlex, titley, fuzz.ratio(str(titlex).strip(), str(titley).strip()))


                        break
                    elif ("," in titlex or ":" in titlex) and ("," in titley or ":" in titley):
                        print("Ne ydalilo: ", titlex, titley, fuzz.ratio(str(titlex).strip(), str(titley).strip()))
                if flag:
                    continue
                break
        if flag:
            continue
        break

    # Видалення дублікатів через rapidfuzz. Кінець функції

    print(f"Статей: {total_stati_count}")
    print(f"Видаленно дублікатів: {removed_stati_count}")
    print(f"Після видалення дублікатів статей: {total_stati_count - removed_stati_count}")


    # Видалення по блеклісту через rapidfuzz. початок функції
    bl_sheet = uploaded_workbook._sheets[5]
    bl_titles = []
    for row in bl_sheet.iter_rows():
        bl_t = row[0].value
        if not bl_t or not str(bl_t).strip():
            continue
        bl_titles.append(str(bl_t).strip())


    total_bl_stati_count = total_stati_count - removed_stati_count
    removed_bl_stati_count = 0
    while True:
        flag = False
        for ix, rowx in enumerate(ws_1.iter_rows()):
            if ix == 0 or rowx[2].value is None:
                continue
            titlex = str(rowx[2].value).strip()

            for _, bl_title in enumerate(bl_titles):
                titley = bl_title
                # ТАК НЕ ПІДХОДИТЬ if fuzz.ratio(str(titlex).strip(), str(titley).strip()) >= 80:

                if (
                        # перша і друга функція # Це удаляє співпадіння с блеклістом які мають рівно 95 або більше процентів схожесті.
                        # Працюють по черзі. Спочатку закоментувати третю. І or теж. Включити бота. Потім закоментувати 1 і 2 функцію. І розкоментувати третю функцію. Включити бота.
                        # Закоментувати спочатку першу і другу одну, включити бота, потім
                        fuzz.ratio(str(titlex).strip(), str(titley).strip()) >= 97  # БУЛО >= 98, 95
                        or
                        fuzz.ratio(" ".join(set(str(titlex).strip().split())), " ".join(set(str(titley).strip().split()))) >= 90  # Було >= 95, потім 40. Удаляє співпадіння з блеклістом в яких слова в заголовці міняються місьцями. Рівно 95 або більше процентів.
                        # третя
                        #or
                        #((("," in titlex or ":" in titlex) and ("," in titley or ":" in titley)) and fuzz.ratio(str(titlex).strip(), str(titley).strip()) < 40)
                ):

                    flag = True
                    ws_1.delete_rows(ix + 1, 1)
                    removed_bl_stati_count += 1

                    print(removed_bl_stati_count, ix, titlex, titley,
                          fuzz.ratio(str(titlex).strip(), str(titley).strip()),
                          fuzz.ratio(" ".join(set(str(titlex).strip().split())), " ".join(set(str(titley).strip().split()))),
                          fuzz.ratio(str(titlex).strip(), str(titley).strip())
                          )

                    break
                # elif ("," in titlex or ":" in titlex) and ("," in titley or ":" in titley):
                #     print("[blacklist] Ne ydalilo: ", titlex, titley,
                #           fuzz.ratio(str(titlex).strip(), str(titley).strip()))

            if flag:
                break

        if flag:
            continue
        break

    # Видалення по блеклісту через rapidfuzz. Кінець функції


    print(f"Видаленно дублікатів по блеклісту: {removed_bl_stati_count}")
    print(f"Після видалення по блеклісту статей: {total_bl_stati_count - removed_bl_stati_count}")



    data_sheet = uploaded_workbook._sheets[0]

    headers = [cell.value for cell in data_sheet[1]]  # Перший рядок містить заголовки
    title_index = headers.index("Title") + 1
    content_index = headers.index("Content") + 1
    url_index = headers.index("URL") + 1

    #Розбити файл ексель по доменах на декілька файлів. Початок функції.

    i_row = 2
    cnt = 0
    domains = set()
    while True:
        end = False

        while i_row <= data_sheet.max_row:
            title = data_sheet.cell(i_row, title_index).value
            url = data_sheet.cell(i_row, url_index).value

            i_row += 1

            # if cnt % 100 == 0:
            #     print("cnt:", cnt, title, url)

            # Допомагає боту зрозуміти де кінець файлу. 100 пустих строк.
            # ПРОБЛЕМА Тестував з 1 сайтом не працює. Додавав +1 сайт +1 дубль, + строки ставив проходити 1, 10, нічого не помогло.
            if cnt >= 100:
                end = True
                break

            if title is None:
                cnt += 1
                continue
            cnt = 0

            domain = url.strip().split("://", 1)[1].split("/")[0]
            domains.add(domain)

        # Була проблема. Ексель раньше часу говорив боту про кінець файлу. І бот не міг найти кінець.
        if end or i_row >= data_sheet.max_row:
            break

    uploaded_workbook.close()

    total_xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
    uploaded_workbook.save(total_xlsx_file_path)

    uploaded_workbook = load_workbook(filename=total_xlsx_file_path)
    data_sheet = uploaded_workbook._sheets[0]

    domains = sorted(list(domains))
    print(domains)
    # input()

    files = []
    for idx, domain in enumerate(domains):
        print(f"Processing domain {domain} ({idx + 1}/{len(domains)})")
        part_xlsx_file_path = f'static/output_{domain}_{random.randint(1000, 9999)}.xlsx'

        with open(total_xlsx_file_path, "rb") as f:
            contents = f.read()
            contents = BytesIO(contents)

            part_workbook = load_workbook(filename=contents)

            data_ws = part_workbook._sheets[0]
            data_ws.delete_rows(2, 2000)

            new_i_row = 2
            old_i_row = 2
            cnt = 0
            while old_i_row <= data_sheet.max_row:
                url = data_sheet.cell(old_i_row, url_index).value

                old_i_row += 1

                if cnt >= 1000:
                    break

                if url is None:
                    cnt += 1
                    continue

                row_url = str(url).strip()

                if domain in row_url:
                    for ic, cell in enumerate(data_sheet[old_i_row - 1]):
                        data_ws.cell(new_i_row, ic + 1, cell.value)
                    new_i_row += 1

            part_workbook.close()
            part_workbook.save(part_xlsx_file_path)

        files.append(part_xlsx_file_path)

    #розбити файл ексель по доменах на декілька файлів. Кінець функції.

    uploaded_workbook.close()

    zip_file_path = f'static/archive_{random.randint(1000, 9999)}.zip'
    create_zip_archive(files, zip_file_path)

    return FileResponse(
        path=zip_file_path,
        media_type='application/zip',
        filename=os.path.basename(zip_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(zip_file_path)}"'
        }
    )



@app.post("/split_by_fix_num")
async def split_by_fix_num(
    file: UploadFile = File(...),
):
    if file == "undefined":
        print("Need file for this operation")
        return HTMLResponse(content="Error: Need file for this operation", status_code=400)

    try:
        # Завантажуємо файл Excel
        contents = await file.read()
        contents = BytesIO(contents)

        # Спробуємо відкрити файл
        uploaded_workbook = load_workbook(filename=contents)
    except:
        print("Can not load excel file")
        return HTMLResponse(content="Error: Can not load excel file", status_code=400)

    print("Start processing data..")

    data_sheet = uploaded_workbook._sheets[0]

    #Розбити файл ексель по 100 статей на декілька файлів. Початок функції.

    i_row = 2
    buff = BytesIO()
    uploaded_workbook.save(buff)
    files = []
    total_rows_count = len([x for x, _ in enumerate(data_sheet.iter_rows())])
    for x in range(total_rows_count):
        if x % 100 == 0:
            part_xlsx_file_path = f'static/output_part_{(x//100)+1}.xlsx'
            files.append(part_xlsx_file_path)

    for x, file_path in enumerate(files):
        print(f"Process file number {x + 1}")
        prev_file = load_workbook(copy(buff))
        p_ws = prev_file._sheets[0]
        p_ws.delete_rows(2, 10000)
        i_row = 2
        for ix, row in enumerate(data_sheet.iter_rows()):
            if ix == 0:
                continue

            if ix < (x*100 + 1) or ix >= ((x+1)*100 + 1):
                continue

            for iy, el in enumerate(row):
                p_ws.cell(i_row, iy + 1, el.value)
            i_row += 1

        prev_file.close()
        prev_file.save(file_path)



    # prev_file = None
    # p_ws = None
    # for ix, row in enumerate(data_sheet.iter_rows()):
    #     if ix == 0:
    #         continue
    #
    #     if (ix - 1) % 100 == 0:
    #         i_row = 2
    #         if prev_file:
    #             prev_file.close()
    #             prev_file.save(files[((ix - 1) // 100)-1])
    #
    #         print(f"Process file number {((ix - 1) // 100)+1}")
    #         prev_file = load_workbook(copy(buff))
    #         p_ws = prev_file._sheets[0]
    #         p_ws.delete_rows(2, 10000)
    #
    #     for iy, el in enumerate(row):
    #         p_ws.cell(i_row, iy+1, el.value)
    #         i_row += 1
    #
    # if prev_file:
    #     prev_file.close()
    #     prev_file.save(files[-1])


    #Розбити файл ексель по 100 статей на декілька файлів. Кінець функції.

    zip_file_path = f'static/archive_{random.randint(1000, 9999)}.zip'
    create_zip_archive(files, zip_file_path)

    return FileResponse(
        path=zip_file_path,
        media_type='application/zip',
        filename=os.path.basename(zip_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(zip_file_path)}"'
        }
    )


@app.post("/concatenate_excel")
async def concatenate_excel(
    file: UploadFile = File(...),
):
    if file == "undefined":
        print("Need file for this operation")
        return HTMLResponse(content="Error: Need file for this operation", status_code=400)

    try:
        # Завантажуємо файл zip
        contents = await file.read()
        contents = BytesIO(contents)

        # Спробуємо відкрити файл
        archive = zipfile.ZipFile(contents)
    except:
        print("Can not load zip file")
        return HTMLResponse(content="Error: Can not load zip file", status_code=400)

    # Склеювання файлів ексель в один. початок функції

    main_workbook = load_workbook(filename=BytesIO(archive.read(archive.infolist()[0].filename)))
    main_workbook._sheets[0].delete_rows(2, 2000)

    main_last_row = 1
    for ix, excel_file in enumerate(archive.infolist()):
        print("Обробка файлу:", excel_file.filename)
        part_workbook = load_workbook(filename=BytesIO(archive.read(excel_file.filename)))

        main_ws = main_workbook._sheets[0]

        part_ws = part_workbook._sheets[0]
        print("Рядків в файлі:", part_ws.max_row)

        for iy, row in enumerate(part_ws.iter_rows()):
            if iy == 0:
                continue

            if not list(filter(lambda x: not (x.value is None), row)):
                continue

            main_last_row += 1
            for iz, cell in enumerate(row):
                main_ws.cell(main_last_row, iz + 1, cell.value)

    # Склеювання файлів ексель в один. Кінець функції

    print(f"Всього файлів склеєно: {len(archive.infolist())}")
    print(f"Всього рядків на виході: {main_last_row}")

    main_workbook.close()

    total_xlsx_file_path = f'static/output_{random.randint(1000, 9999)}.xlsx'
    main_workbook.save(total_xlsx_file_path)

    return FileResponse(
        path=total_xlsx_file_path,
        media_type='application/xlsx',
        filename=os.path.basename(total_xlsx_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(total_xlsx_file_path)}"'
        }
    )



@app.post("/fetch_urls_from_xml")
async def fetch_urls_from_xml(request: Request):
    data = await request.form()
    print(data)

    file = data.get('file')
    wordsList = data.get('wordsList')

    if not file or file == "undefined":
        print("Need file for this operation")
        return HTMLResponse(content="Error: Need file for this operation", status_code=400)

    if not wordsList or wordsList == "undefined":
        print("Need words list for this operation")
        return HTMLResponse(content="Error: Need words list for this operation", status_code=400)

    try:
        # Завантажуємо файл zip
        contents = await file.read()
        contents = BytesIO(contents)

        # Спробуємо відкрити файл
        xml_text = contents.getvalue().decode("utf-8")
    except:
        print("Can not load file")
        return HTMLResponse(content="Error: Can not load file", status_code=400)

    # Діставання урлів з хмл. початок функції

    url_extract_pattern = "https?:\\/\\/(?:www\\.)?[-a-zA-Z0-9@:%._\\+~#=]{1,256}\\.[a-zA-Z0-9()]{1,6}\\b(?:[-a-zA-Z0-9()@:%_\\+.~#?&\\/=]*)"

    finds = re.findall(url_extract_pattern, xml_text)

    words_list = wordsList.split(",")
    words_list = [x.strip() for x in words_list]

    print(f"Всього посилань (не фільтрованих): {len(finds)}")

    result_links = []
    for url in finds:
        for w in words_list:
            if w in url:
                result_links.append(url)
                break

    print(f"Всього рядків на виході: {len(result_links)}")

    # Діставання урлів з хмл. Кінець функції

    result_links_file_path = f'static/result_links_{random.randint(1000, 9999)}.txt'
    with open(result_links_file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(result_links))

    return FileResponse(
        path=result_links_file_path,
        media_type='application/txt',
        filename=os.path.basename(result_links_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(result_links_file_path)}"'
        }
    )


@app.get("/downloadPicture", response_class=HTMLResponse)
async def downloadPicture(request: Request):
    return period_check(templates.TemplateResponse("downloadPicture.html", {"request": request}))

@app.post("/save_duplicate")
async def save_to_duplicate(updated_contents: UpdatedContent):
    try:
        sheet = uploaded_workbook.active
        # Отримуємо заголовки стовпців
        headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки

        # Індекси потрібних стовпців
        content_index = headers.index("Content")+1
        i_row = 2
        while i_row <= sheet.max_row:
            content = sheet.cell(i_row, content_index)  # Беремо тільки "Content"
            # Перевіряємо, чи є контент та на дублікат
            if content.value:
                try:
                    duplicate_clear.add_content(content.value)
                except UnicodeEncodeError:
                    print("Error: збереження дублікату\n", content.value)
            i_row += 1
        return JSONResponse(status_code=200, content=dict(status="ok"))
    except duplicate_clear.BlockDuplicateFile:
        return JSONResponse(status_code=400, content={"error": "Файл дублікатів заблокований іншою програмою"})


@app.get("/download_docx/{file_name}")
async def download_docx(file_name: str):
    """
    Видає Word файл на скачання.
    Anthor: @Teri_anric
    """
    return FileResponse(
            file_name,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            content_disposition_type="attachment",
            filename=file_name
        )

@app.post("/save_docx")
async def save_docx(updated_contents: UpdatedContent):
    """
    Генерує zip файл (в якому файли Word) статтів.
    Видає посилання на завантаження (функція download_docx)
    Логілка описана в фалі html_to_docx
    Anthor: @Teri_anric
    """
    docs = []
    for i, data in enumerate(updated_contents.updated_contents, start=1):
        output = f"Візуальний редактор {i}.docx"
        generate_document(data["content"], output)
        docs.append(output)
    zip_filename = "Статті.zip"
    create_zip_archive(docs, zip_filename)
    for filename in docs:
        os.unlink(filename)
    return {"files": [f"/download_docx/{zip_filename}"]}



if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)







