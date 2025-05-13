from fastapi.staticfiles import StaticFiles
import asyncio
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
import os
import random
import time
import yaml
from urllib.parse import unquote, urlparse
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import aiohttp
import logging
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import shutil
from io import BytesIO
from html import unescape
from playsound import playsound

# Глобальні змінні для зберігання статистики завантажень
request_success_count = 0
request_failure_count = 0
selenium_success_count = 0
selenium_failure_count = 0
failure_status_codes = {}


# Налаштування логів
logging.basicConfig(level=logging.INFO, filename='parser.log', filemode='a', format='%(asctime)s - %(levelname)s - %(message)s')

#Генерація випадкового тексту
def gen_rand_text(text_len: int = None):
    alph = list("abcdefghijklmnopqrstuvwxyz")
    random.shuffle(alph)
    return "".join(alph[:(text_len or 5)])
claimed_names = []

# Функція для виведення статистики
def log_final_statistics():
    global request_success_count, request_failure_count, selenium_success_count, selenium_failure_count, failure_status_codes

    total_success = request_success_count + selenium_success_count
    total_failures = sum(failure_status_codes.values()) + selenium_failure_count

    if failure_status_codes:
        failure_codes_list = [f"{code}: {count}" for code, count in failure_status_codes.items()]
        print(f"Через помилки: {', '.join(failure_codes_list)}")  # Додано для перевірки
        logging.info(f"Через помилки: {', '.join(failure_codes_list)}")
    else:
        print("Не було зафіксовано жодних статус-кодів помилок.")
        logging.info("Не було зафіксовано жодних статус-кодів помилок.")

    # Завжди виводимо загальну статистику
    print(f"СКАЧАВ: {total_success}")  # Додано для перевірки
    print(f"НЕ СКАЧАВ: {total_failures}")  # Додано для перевірки

    logging.info(f"СКАЧАВ: {total_success}")
    logging.info(f"НЕ СКАЧАВ: {total_failures}")


# Завантаження конфігураційного файлу
def load_config(config_file: str):
    """
    Завантажує конфігурацію з YAML файлу.
    """
    if not os.path.isfile(config_file):
        raise FileNotFoundError(f"Файл конфігурації {config_file} не знайдено.")
    with open(config_file, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

# Отримання проксі з конфігураційного файлу
def get_proxy():
    if not 1: # ВИМКНУТИ ПРОКСІ ЯКИЙ ПАРСИТЬ САЙТИ НА РЕКВЕСТАХ. 1 - юзати проксі, 0 - не юзати
        return
    config = load_config("config.yaml")
    PROXY_CHECK_URL = config.get("proxy_check_url", "https://ya.ru/")
    if not config.get("proxy_config", []):
        return
    proxy_list = config.get('proxy_config', [])
    if not isinstance(proxy_list, list):
        proxy_list = [proxy_list]

    while proxy_list:
        proxy = random.choice(proxy_list)
        p_url = urlparse(proxy)
        try:
            proxies = {
                "http": proxy,
                "https": proxy
            }
            if p_url.scheme in ["socks5", "socks6"]:
                proxies = { p_url.scheme: proxy }

            r = requests.get(PROXY_CHECK_URL, proxies=proxies)
            r.raise_for_status()
            print(f"Використано проксі-сервер({proxy}): {r.text[:50]}")
            return proxy
        except Exception as e:
            print("Проксі не пройшов перевірку ", proxy, e)
            proxy_list.remove(proxy)
    print(f"Усі проксі не пройшли перевірку\n МОЖЛИВО НЕ ПРАЦЮЄ САЙТ {PROXY_CHECK_URL}")

# Скачування/завантаження картинок. Початок коду.
# temp name

# Налаштування статичних файлів
static_dir = 'templates'
os.makedirs(static_dir, exist_ok=True)
os.makedirs('static', exist_ok=True)
os.makedirs('templates/static', exist_ok=True)
import os
if not os.path.exists("templates/static/parsed_result.css"):
    print("CSS файл не знайдено!")
app = FastAPI()

# Підключення статичних файлів (якщо ще не додано)
from fastapi.staticfiles import StaticFiles
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="templates/static"), name="static")

@app.get("/", response_class=HTMLResponse)
async def root():
    """Відображення сторінки downloadPicture.html"""
    try:
        with open("templates/downloadPicture.html", "r", encoding="utf-8") as file:
            return file.read()
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Файл downloadPicture.html не знайдено</h1>", status_code=404)

# Скачування/завантаження картинок. Початок коду.
@app.post("/process_excel_2")
async def process_excel_2(request: Request):
    # Ініціалізація списку claimed_names
    claimed_names = []

    data = await request.form()

    if os.path.exists('static/results'):
        shutil.rmtree('static/results')

    if not os.path.exists('static/results'):
        os.mkdir('static/results')

    folder_name = data.get('folderName')
    file = data.get('file')

    if folder_name and folder_name.strip():
        try:
            safe_title = re.sub(r'[\/\\:*?"<>|\t\n]', '_', folder_name).strip()
            download_folder = f'static/results/{safe_title}'

            if not os.path.exists(f'static/results/{safe_title}'):
                os.mkdir(f'static/results/{safe_title}')
        except Exception as e:
            logging.error(f'Назва має некоректний символ: {e}')
            return HTMLResponse(content="<h1>Error: Invalid folder name</h1>", status_code=400)
    else:
        if not os.path.exists('static/results/image'):
            os.mkdir('static/results/image')
        download_folder = 'static/results/image'

    # Завантажуємо файл Excel
    contents = await file.read()

    # Спробуємо відкрити файл
    uploaded_workbook = load_workbook(filename=BytesIO(contents))

    # Отримання даних з першого листа
    sheet = uploaded_workbook.active

    # Отримуємо заголовки стовпців
    headers = [cell.value for cell in sheet[1]]  # Перший рядок містить заголовки

    # Індекси потрібних стовпців
    content_index = headers.index("Content")+1
    orig_urls_index = headers.index("Image Url_original")+1
    image_now_urls_index = headers.index("Image now Url")+1

    # https://sykni.fff/wp-content/uploads/2024/09/                                              bearded_european_man_in_casual_peach_isolated_excited_cheerful_holding_gift_box.jpg
    # https://ulanude.royalthai.ru/upload/medialibrary/ef3/qqtyncldrg4wsrdj3pesahoyyx4ezt98/     bearded_european_man_in_casual_peach_isolated_excited_cheerful_holding_gift_box.jpg

    i_row = 2
    cnt = 0

    # ЗВУК ПРИ ПОЧАТКУ ПАРСИНГУ ЩОБ РОЗКОМЕНТУВАТИ, ПОТРІБНО УДАЛИТИ #
    playsound(r'./audio/rozpochala_parsing.mp3')
    print("Завантаження розпочато...")

    import requests
    session = requests.session()

    while i_row <= sheet.max_row:
        content = sheet.cell(i_row, content_index).value  # Беремо тільки "Content"
        orig_urls = sheet.cell(i_row, orig_urls_index).value  # Беремо тільки "Image Url_original"

        sheet.cell(i_row, image_now_urls_index, "")

        if cnt >= 1000:
            break

        if not orig_urls:
            cnt += 1
            continue

        new_links = []

        print(f"Processing line {i_row}..")

        soup = BeautifulSoup(content, "html.parser")

        for url in list(set(orig_urls.split(" "))):
            # print(url)
            url = url.split("?")[0]
            need_clear = False
            pic_paths = []
            if not "." in url[-6:]: # шука точку в останніх 6 символах ссилки
                need_clear = True
            else:
                pic_paths, success_count, error_count = await download_images_v2([url], download_folder, session)
                # Оновлення глобальних змінних
                global request_success_count, request_failure_count
                request_success_count += success_count
                request_failure_count += error_count
                if not pic_paths:
                    need_clear = True
            # print(pic_paths)

            # Отримання імені файлу з URL
            parsed_url = urlparse(url)
            file_name = os.path.basename(unquote(parsed_url.path))
            renamed = False

            if not need_clear:
                name, file_format = file_name.rsplit(".", 1)
                while True:
                    new_file_name = gen_rand_text() + "." + file_format
                    if new_file_name in claimed_names:
                        continue
                    claimed_names.append(new_file_name)
                    if not pic_paths:
                        print('pic_paths', pic_paths)
                        break
                    if not pic_paths[0]:
                        print('pic_paths[0]', pic_paths)
                        break
                    if pic_paths[0] == []:
                        print('pic_paths[0]==[]', pic_paths)
                        break
                    if isinstance(pic_paths[0], str):
                        os.rename(pic_paths[0], download_folder + "/" + new_file_name)
                    if isinstance(pic_paths[0], list):
                        os.rename(pic_paths[0][0], download_folder + "/" + new_file_name)
                    break

            # Process exactly image tags
            for html_img_tag in soup.find_all("img"):
                try:
                    html_img_url = html_img_tag["src"]
                    if not html_img_url:
                        raise ValueError()
                except:
                    # Remove img tag if no src
                    print(f"No src in img tag {html_img_tag}")
                    if html_img_tag.parent.name == "p":
                        try:
                            html_img_tag.parent.replace_with("")
                        except ValueError:
                            newtext = re.sub(unescape(str(html_img_tag.parent)), '', unescape(str(soup)))
                            soup = BeautifulSoup(newtext, "html.parser")
                    else:
                        try:
                            html_img_tag.replace_with("")
                        except ValueError:
                            newtext = re.sub(unescape(str(html_img_tag)), '', unescape(str(soup)))
                            soup = BeautifulSoup(newtext, "html.parser")

                    continue
                # print(html_img_tag, html_img_url, unquote(html_img_url))

                if need_clear:
                    # Find bad image
                    if file_name in unquote(html_img_url):
                        # Remove link from html
                        if html_img_tag.parent.name == "p":
                            # print("p", html_img_tag.parent, html_img_tag.parent.parent)
                            try:
                                html_img_tag.parent.replace_with("")
                            except ValueError:
                                # print(html_img_tag.parent, html_img_tag.parent.parent)
                                newtext = re.sub(unescape(str(html_img_tag.parent)), '', unescape(str(soup)))
                                soup = BeautifulSoup(newtext, "html.parser")
                        else:
                            # print("not p", html_img_tag)
                            try:
                                html_img_tag.replace_with("")
                            except ValueError:
                                newtext = re.sub(unescape(str(html_img_tag)), '', unescape(str(soup)))
                                soup = BeautifulSoup(newtext, "html.parser")
                        # print(parsed_url, file_name, html_img_url)
                    continue

                # Anyway rename
                if file_name in unquote(html_img_url):
                    link_start = unquote(html_img_url).split(file_name)[0]
                    link_new = link_start + new_file_name
                    html_img_tag["src"] = link_new

                    nl = html_img_tag["src"]
                    new_links.append(nl)
                    # print(f"fn {file_name}")
                    # print(f"start {link_start}")
                    # print(f"nfn {new_file_name}")
                    # print(f"old src {html_img_url}")
                    # print(f"new src {nl}")
                    renamed = True

            # Clear other tags from bad url
            for html_img_tag in soup.find_all():
                if file_name in unquote(str(html_img_tag)):
                    new_tag = unquote(str(html_img_tag))
                    new_tag = new_tag[:new_tag.rfind("\"", 0, new_tag.find(file_name)) + 1] + new_tag[new_tag.find("\"", new_tag.find(file_name)):]
                    try:
                        html_img_tag.replace_with(new_tag)
                    except ValueError:
                        newtext = re.sub(unescape(str(html_img_tag)), new_tag, unescape(str(soup)))
                        soup = BeautifulSoup(newtext, "html.parser")

            if not need_clear and not renamed:
                print(f"Can not replace photo in excel (bo nema tega src), additional info: file_name {file_name}, new_file_name {new_file_name}, pic_paths {pic_paths}, url {url}")
                try:
                    os.remove(download_folder + "/" + new_file_name)
                    print(f"Photo file deleted")
                except:
                    print(f"[Error] Can not delete photo file on path {download_folder}/{new_file_name}")

            # print(len(list(soup.find_all("img"))))

            sheet.cell(i_row, content_index, unescape(str(soup)))

        if new_links:
            sheet.cell(i_row, image_now_urls_index, " ".join(new_links))

        i_row += 1


    # ЗВУК ПРИ КІНЦІ ПАРСИНГУ
    playsound(r'./audio/zaversheno_parsing.mp3')
    print("Завантаження завершено")
# Скачування/завантаження картинок. Кінец функції завантаження.

    # zip_file_path = f'static/results/archive_{random.randint(1000, 9999)}'
    # zip_filename = shutil.make_archive(zip_file_path, 'zip', 'static/v2')

    xlsx_file_path = f'{download_folder}/table_{random.randint(1000, 9999)}.xlsx'
    uploaded_workbook.save(xlsx_file_path)

    end_zip_file_path = f'static/archive_{random.randint(1000, 9999)}'
    zip_filename = shutil.make_archive(end_zip_file_path, 'zip', download_folder)

    # Виведення фінальної статистики після завершення роботи
    print("Статистика завантаження картинок.")
    # Додано перевірку глобальних змінних
    #print(f"Скачано через реквести: {request_success_count}") # не рахує правильно, все зараховує реквестам, навіть те що скачав селеніум
    #print(f"request_failure_count: {request_failure_count}")
    #print(f"Скачано через selenium: {selenium_success_count}") не працює
    print(f"Selenium не скачав: {selenium_failure_count}")
    print(f"Коди відповіді і їх кількість, через які не скачалось: {failure_status_codes}")
    log_final_statistics()

    return FileResponse(
        path=end_zip_file_path + '.zip',
        media_type='application/zip',
        filename=os.path.basename(end_zip_file_path),
        headers={
            'Content-Disposition': f'attachment; filename="{os.path.basename(end_zip_file_path)}"'
        }
    )
# Скачування/завантаження картинок. Кінець коду.


# Скачування Selenium Ware при помилках 503, 403
def selenium_download_image(proxy_url, file_url, download_folder):
    proxy_options = {
        'proxy': {
            'http': proxy_url,
            'https': proxy_url,
            'no_proxy': 'localhost,127.0.0.1'
        }
    }

    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    chrome_options.add_argument("--disable-logging")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument("--enable-unsafe-swiftshader")
    chrome_options.add_argument("--disable-software-rasterizer")


    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=chrome_options,
        seleniumwire_options=proxy_options
    )

    try:
        driver.get("https://api.ipify.org")
        current_ip = driver.find_element(By.TAG_NAME, "body").text
        # time.sleep(2)
        print('IP: ', current_ip)

        driver.get(file_url)
        time.sleep(random.uniform(2, 8))

        if "captcha" in driver.page_source.lower():
            input(f"Обнаружена капча! Решите её и нажмите Enter для продолжения...")

        img_element = driver.find_element(By.TAG_NAME, "img")

        if not os.path.exists(download_folder):
            os.makedirs(download_folder)

        parsed_url = urlparse(file_url)
        original_filename = os.path.basename(parsed_url.path)
        if not original_filename:
            original_filename = f"downloaded_{int(time.time())}.png"
        else:
            name, ext = os.path.splitext(original_filename)
            original_filename = f"{name}{ext}"

        save_path = os.path.join(download_folder, original_filename)

        img_element.screenshot(save_path)
        print(f"Изображение сохранено в {save_path}")
        return True
    except Exception as e:
        global selenium_failure_count
        selenium_failure_count += 1
        print(f"Ошибка: {e}")
        return False
    finally:
        if selenium_success_count > 0:
            logging.info(f"[Selenium] Успішно завантажено зображень: {selenium_success_count}")
        if selenium_failure_count > 0:
            logging.info(f"[Selenium] Не вдалося завантажити зображень: {selenium_failure_count}")
        driver.quit()

# Получаем ip через aiohttp
async def fetch_ip(proxy=None):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get("https://api.ipify.org", proxy=proxy) as response:
                ip = await response.text()
                return ip
    except Exception as e:
        print(f"[aiohttp] Ошибка при получении IP: {e}")
        return None

# -------------------------------------------------------
def get_status_description(status_code: int)->str:
    """Повертає опис статус-коду HTTP."""
    from http import HTTPStatus
    try:
        description = HTTPStatus(status_code).description
        return f"Код відповіді: {status_code} - {description}"
    except ValueError:
        return f"Невідомий код відповіді: {status_code}"

#  Скачування картинок реквестами
async def download_images_v2(image_urls, download_folder, session = None):
    global request_success_count, request_failure_count, failure_status_codes, selenium_success_count, selenium_failure_count, status_description
    if not os.path.exists(download_folder):
        os.makedirs(download_folder)
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7',
        'priority': 'u=0, i',
        #'referer': 'https://www.eldorado.ru/',
        'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Google Chrome";v="134"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Linux"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-site',
        'upgrade-insecure-requests': '1',
        #'user-agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36',
        'user-agent': 'Mozilla/5.0 AppleWebKit/537.36 (KHTML, like Gecko; compatible; Googlebot/2.1; +http://www.google.com/bot.html) Chrome/118.0.5993.70 Safari/537.36'
    }

    # Заміна URL як що силка редеріктить.
    #image_urls = [url.replace("https://www.eldorado.ru", "https://static.eldorado.ru") for url in image_urls]

    proxy = None
    # Проксі для скачування картинок
    # закоментувати proxy = get_proxy() щоб виключити проксі саме в цій функції

    list_image_path = []
    if not isinstance(image_urls, list):
        image_urls = [image_urls]
    #proxy = get_proxy()
    # old code
    """
    retry_count = 0
    max_retries = 5
    retry_sleep = 3
    between_urls_sleep = 3
    for i, url in enumerate(image_urls):
        while retry_count < max_retries:
            retry_count += 1
            try:
                response = session.get(url, headers=headers, verify=False)
                # if response.status_code in [429, 403, 503, 502, 500]:
                #     # use selenium
                #     from parser import Selenium_Parser # local import to avoid circle imports
                #     selenium_result = Selenium_Parser(url)

                response.raise_for_status()  # Перевірка на помилки

                # Отримання імені файлу з URL
                parsed_url = urlparse(url)
                file_name = os.path.basename(unquote(parsed_url.path))
                name, file_format = file_name.rsplit(".", 1)
                while True:
                    new_file_name = gen_rand_text() + "." + file_format
                    if new_file_name in claimed_names:
                        continue
                    claimed_names.append(new_file_name)
                    # file_path = os.path.join(download_folder, new_file_name) # file_name
                    file_path = download_folder + "/" + new_file_name # file_name
                    break

                with open(file_path, 'wb') as file:
                    file.write(response.content)

                list_image_path.append(file_path)
            except Exception as e:
                print(f"Error downloading {url}: {e}")
                logging.info(f"Помилка: URL '{url[:15]}' не є вірним або сталася інша помилка під час завантаження.")
                if retry_count < max_retries:
                    await asyncio.sleep(retry_sleep)
        await asyncio.sleep(between_urls_sleep)
    """

    errorDownload = 0

    for i, url in enumerate(image_urls):
        retry_count = 0
        max_retries = 5

        while retry_count < max_retries:
            proxy = get_proxy()
            print(f"\n{'=' * 100}")
            print(f"[HTTPS][СПРОБА {retry_count + 1}/{max_retries}] {url}")

            random_delay = random.uniform(2, 5)
            print(f"[HTTPS] Очікування {random_delay:.1f}с...")
            await asyncio.sleep(random_delay)

            try:
                timeout_time = 15
                timeout = aiohttp.ClientTimeout(total=timeout_time)

                async with aiohttp.ClientSession(timeout=timeout) as session:
                    async with session.get(url, headers=headers, proxy=proxy) as response:
                        status_code = response.status

                        ip_address = await fetch_ip(proxy)
                        if ip_address:
                            print("Текущий IP (aiohttp):", ip_address)
                        else:
                            print("Не удалось получить IP через aiohttp.")

                        global status_description
                        status_description = get_status_description(response.status)

                        # Якщо статус-код 404, припиняємо спроби
                        if status_code == 404:
                            print(f"[HTTPS] Помилка {status_code} - {status_description}. Припиняємо спроби для {url}.")
                            logging.warning(f"[HTTPS] Помилка {status_code} - {status_description} для {url}.")
                            if status_code not in failure_status_codes:  # Уникнення дублювання
                                failure_status_codes[status_code] = 0
                            failure_status_codes[status_code] += 1
                            #print(f"[DEBUG] request_failure_count до збільшення: {request_failure_count}")
                            request_failure_count += 1  # Збільшуємо лише тут
                            #print(f"[DEBUG] request_failure_count після збільшення: {request_failure_count}")
                            errorDownload += 1
                            break  # Виходимо з циклу спроб

                        # Якщо статус-код 429, 502 або 500, робимо повторну спробу із затримкою
                        if status_code in [429, 502, 500]:
                            wait_time = random.uniform(5, 15)
                            print(f"[HTTPS] Помилка {status_code}, {status_description}. Очікування {wait_time:.1f}с")
                            logging.warning(f"Код {status_code}: {status_description} для {url}")
                            retry_count += 1  # Лише збільшуємо лічильник спроб
                            await asyncio.sleep(wait_time)
                            continue
                            # Додати код помилки: Щоб селеніум качав.
                        if status_code in [403, 503, 500]:
                           print(f"[HTTPS] Помилка {status_code} --> Selenium")
                           statusS = selenium_download_image(proxy, url, download_folder)
                           if statusS:
                               parsed_url = urlparse(url)
                               file_name = os.path.basename(unquote(parsed_url.path))
                               try:
                                   name, file_format = file_name.rsplit(".", 1)
                               except ValueError:
                                   # Если расширение не удалось определить, задаём стандартное
                                   file_format = "jpg"
                               new_file_name = gen_rand_text() + "." + file_format
                               file_path = os.path.join(download_folder, file_name)
                               new_file_path = os.path.join(download_folder, new_file_name)
                               if os.path.exists(file_path):
                                   os.rename(file_path, new_file_path)
                                   list_image_path.append(new_file_path)
                                   print(f"[Selenium] Успішно")
                               else:
                                   errorDownload += 1
                                   print(f"[Selenium] Error")
                           break

                        if status_code == 200:
                            print(f"[HTTPS] Успішно")
                            print(f"{'=' * 100}")
                            # Отримання імені файлу з URL
                            parsed_url = urlparse(url)
                            file_name = os.path.basename(unquote(parsed_url.path))
                            name, file_format = file_name.rsplit(".", 1)
                            while True:
                                new_file_name = gen_rand_text() + "." + file_format
                                if new_file_name in claimed_names:
                                    continue
                                claimed_names.append(new_file_name)
                                file_path = os.path.join(download_folder, new_file_name)
                                break

                            with open(file_path, "wb") as f:
                                f.write(await response.read())

                            list_image_path.append(file_path)
                            print(f"[INFO] Додано новий файл до списку: {file_path}")
                            break  # Завершуємо спроби, якщо успішно
            except (asyncio.TimeoutError, Exception) as e:
                print(f"[HTTPS] Помилка: {str(e)}")
                retry_count += 1  # Збільшуємо лише лічильник спроб
                await asyncio.sleep(random.uniform(3, 7))
            else:
                print(f"[HTTPS] Помилка {status_code}")
                retry_count += 1
                if retry_count == max_retries:
                    errorDownload += 1
                    # Оновлення глобальних змінних
                    print(f"[DEBUG] request_failure_count до збільшення: {request_failure_count}")
                    request_failure_count += 1
                    print(f"[DEBUG] request_failure_count після збільшення: {request_failure_count}")
                    failure_status_codes[status_code] = failure_status_codes.get(status_code, 0) + 1
                await asyncio.sleep(random.uniform(3, 7))

    logging.info(f"[Реквести] Успішно завантажено зображень: {request_success_count}")
    logging.info(f"[Реквести] Не вдалося завантажити зображень: {request_failure_count}")
    logging.info(f"[Реквести] Коди помилок: {failure_status_codes}")
    return list_image_path, len(list_image_path), errorDownload
